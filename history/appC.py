import logging
import os
import re
from collections import deque
from datetime import datetime
from pprint import pprint
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
import urllib3
from bs4 import BeautifulSoup, UnicodeDammit
from openpyxl.styles import Alignment, Font, PatternFill
from ruamel.yaml import YAML

# from openpyxl.worksheet.hyperlink import Hyperlink
# https://titangene.github.io/article/python-logging.html


def is_valid_link(base_url, link) -> bool:
    '''檢查連結是否有效'''
    global visited_link
    full_url = urljoin(base_url, link).strip()  # 取得絕對連結, 並去除前後空白
    if full_url in AVOID_URLS:  # 若連結為避免檢查的網址
        status = "200 (避免檢查的網址)"
    elif full_url in visited_link:  # 若連結已經檢查過
        status = visited_link[full_url]  # 取得狀態碼
    else:
        protocol = urlparse(full_url).scheme  # 取得協定
        try:
            print(f"檢查: {link} ... ", flush=True, end="\r")  # 顯示目前檢查的連結，不快取強制輸出後不換行
            if protocol == "https":
                response = requests.get(
                    full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True
                )  # 發送 GET 請求，其實發 HEAD 請求會比較好，但會返回 403 很奇怪，可能是 Server 端設定錯誤
            else:
                response = requests.get(full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, verify=False)
            status = str(response.status_code)  # 取得狀態碼
            real_url = response.url  # 取得網頁的實際連結, 避免重定向造成誤判
            domain1 = urlparse(full_url).netloc  # 取得原始連結的網域
            domain2 = urlparse(real_url).netloc  # 取得實際連結的網域
            # domain1 = '.'.join(urlparse(full_url).netloc.split('.')[1:])  # 取得原始連結的父網域
            # domain2 = '.'.join(urlparse(real_url).netloc.split('.')[1:])  # 取得實際連結的父網域
            if status == '200' and domain1 != domain2:  # 若原始連結與實際連結的網域不同(例如: 302 暫時定向到別的網域)
                if protocol == "https":
                    response = requests.get(full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=False)
                else:
                    response = requests.get(
                        full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=False, verify=False
                    )
                status = f'200 重定向：{str(response.status_code)} 到 {real_url}'  # 設定狀態
        except requests.exceptions.ConnectTimeout as e:
            status = f"連線逾時：{link}  錯誤訊息：{e}"
        except requests.exceptions.ConnectionError as e:
            status = f"無法連線至此網頁：{link}  錯誤訊息：{e}"
        except requests.exceptions.Timeout as e:
            status = f"連線逾時：{link}  錯誤訊息：{e}"
        except requests.exceptions.MissingSchema as e:
            status = f"連結缺少協定：{link}  錯誤訊息：{e}"
        except requests.exceptions.TooManyRedirects as e:
            status = f"重新導向次數過多：{link}  錯誤訊息：{e}"
        except requests.exceptions.SSLError as e:
            status = f"SSL 錯誤或憑證不正確：{link}  錯誤訊息：{e}"
        except requests.exceptions.RequestException as e:
            status = f"無法取得此網頁內容：{link}  錯誤訊息：{e}"
        except requests.exceptions.HTTPError as e:
            status = f"HTTP 錯誤：{link}  錯誤訊息：{e}"
        except Exception as e:
            status = f"其它錯誤：{link}  錯誤訊息：{e}"
        if status == "403":
            status = f"{status} 請求被拒絕"
        elif status == "404":
            status = f"{status} 請求失敗"
        elif protocol == "http":
            status = f"{status} 且使用 http 協定並不安全"
        visited_link[full_url] = status  # 將連結與狀態碼加入已檢查的集合

    if "200" in status:
        logger.info(f"檢查: {link} ... 狀態: {status}")
    else:
        logger.error(f"檢查: {link} ... 狀態: {status}")
    return status


def get_links(url) -> tuple:
    '''取得網頁中的所有內部連結、外部連結與沒有 alt 的連結'''
    try:
        protocol = urlparse(url).scheme  # 取得協定
        domain = urlparse(url).netloc  # 取得網域
        if protocol == "https":
            response = requests.get(
                url.strip(), headers=HEADERS, timeout=TIMEOUT, allow_redirects=True
            )  # 發送 GET 請求
        else:
            response = requests.get(url.strip(), headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, verify=False)
        real_url = response.url  # 取得網頁的實際連結, 避免重定向造成誤判
        real_domain = urlparse(real_url).netloc  # 取得網域
        if domain != real_domain:  # 若原始連結與實際連結的網域不同
            return ([], [], [])  # 回傳空串列
        soup = BeautifulSoup(
            UnicodeDammit(response.content, ["latin-1", "iso-8859-1", "windows-1251"]).unicode_markup, 'lxml'
        )  # 使用 BeautifulSoup 解析網頁內容
        result = soup.find("meta", attrs={"http-equiv": "refresh"})  # 取得 refresh 標籤
        if result:  # 若有 refresh 標籤，代表網頁有前端重新導向
            wait, client_url = result["content"].split(";")  # 取得重新導向的秒數與網址
            if client_url.strip().lower().startswith("url="):  # 若網址以 url= 開頭
                client_url = urljoin(url, client_url.strip()[4:])  # 取得絕對網址
            if protocol == "https":
                response = requests.get(
                    client_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True
                )  # 發送 GET 請求
            else:
                response = requests.get(
                    client_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, verify=False
                )
            soup = BeautifulSoup(
                UnicodeDammit(response.content, ["latin-1", "iso-8859-1", "windows-1251"]).unicode_markup, 'lxml'
            )  # 使用 BeautifulSoup 解析網頁內容
        all_links = []  # 取得所有的連結
        internal_links, external_links, no_alt_links = [], [], []  # 初始化內部連結、外部連結與沒有 alt 的連結

        hrefs = [tag.get('href').strip() for tag in soup.find_all(href=True)]  # 找到所有具有 href 屬性的屬性值
        srcs = [tag.get('src').strip() for tag in soup.find_all(src=True)]  # 找到所有具有 src 屬性的屬性值
        if ALT_MUST:  # 若須偵測圖片的 alt 屬性
            no_alt_links = [
                tag.get('src') for tag in soup.find_all('img', alt=False)
            ]  # 找到所有圖片沒有 alt 屬性的標籤
        all_links.extend(hrefs)  # 合併 href 屬性值到所有連結
        all_links.extend(srcs)  # 合併 src 屬性值到所有連結
        all_links = list(set(all_links))  # 去除重複的連結
        for link in all_links:
            link_domain = urlparse(link).netloc  # 取得連結的網域
            if link.startswith(
                ('#', 'javascript', 'mailto', 'data:image')
            ):  # 若連結為錨點、javascript、郵件連結或 base64圖片，則跳過
                continue
            elif link_domain == domain:  # 若連結為目前網址
                internal_links.append(link)
            elif link.startswith(('http', '//')):  # 若連結為 http 或 // 開頭的為外部連結
                external_links.append(link)
            elif link.startswith('/'):  # 若連結為 / 開頭的為內部連結
                internal_links.append(link)
            else:  # 其它的為內部連結，例如檔名開頭的
                internal_links.append(link)
        return (internal_links, external_links, no_alt_links)  # 回傳內部連結、外部連結、圖片沒有 alt 屬性的連結
    except requests.RequestException as e:
        logger.error(f"無法取得此網頁內容：{url}  錯誤訊息：{e}")
        return ([], [], [])  # 若發生錯誤，則回傳空串列


def queued_link_check(start_url, depth_limit=1) -> list:
    '''使用雙向佇列結構儲存網站中的連結'''
    visited_url = set()  # 存放已經檢查過的連結
    queue = deque([(start_url, 0)])  # 存放待檢查的連結，設定目前深度為 0
    all_err_links = []  # 存放所有錯誤連結
    while queue:
        url, current_depth = queue.popleft()  # 取得待檢查的連結
        if url in AVOID_URLS:  # 若連結為避免檢查的網址
            continue  # 跳過
        match = re.search(r".*\.\w{2,4}$", url)  # 是否符合具有副檔名的.xxx或.xxxx (例如: .7z, .php, .aspx)
        if match and not url.endswith(
            ('html', 'htm', 'php', 'asp', 'aspx', 'jsp', 'tw', 'com')
        ):  # 若連結為檔案且非網頁檔
            continue  # 跳過
        if url not in visited_url and current_depth <= depth_limit:  # 若連結未檢查過且深度未達到指定深度
            logger.info(f"第 {current_depth} 層連結： {url}")  # 顯示目前檢查的連結
            visited_url.add(url)  # 將連結加入已檢查的集合
            internal_links, external_links, no_alt_links = [], [], []  # 初始化內部連結、外部連結與沒有 alt 的連結
            internal_links, external_links, no_alt_links = get_links(url)  # 取得內部連結、外部連結與沒有alt的連結
            if internal_links:
                logger.info("內部連結：" + ', '.join(internal_links))  # 顯示內部連結
            if external_links:
                logger.info("外部連結：" + ', '.join(external_links))  # 顯示外部連結
            if no_alt_links:
                logger.info("沒有 alt 屬性的連結：" + ', '.join(no_alt_links))  # 顯示沒有 alt 屬性的連結

            error_internal_links = []  # 存放錯誤的內部連結
            for link in internal_links:
                status = is_valid_link(url, link)  # 檢查連結是否有效
                if '200' not in status:  # 若狀態碼不是 200
                    error_internal_links.append((link, status))  # 將錯誤的連結加入存放錯誤的內部連結
            error_external_links = []  # 存放錯誤的外部連結
            for link in external_links:  # 檢查外部連結是否有效
                status = is_valid_link(url, link)  # 檢查連結是否有效
                if '200' not in status:  # 若狀態碼不是 200
                    error_external_links.append((link, status))  # 將錯誤的連結加入存放錯誤的外部連結

            error_links = []  # 存放此次所有 url 錯誤的連結
            # 將錯誤的內部連結加入所有錯誤連結集合
            error_links.extend(error_internal_links) if error_internal_links else error_links
            # 將錯誤的外部連結加入所有錯誤連結集合
            error_links.extend(error_external_links) if error_external_links else error_links
            error_links = list(set(error_links))  # 去除重複的錯誤連結
            if error_links:  # 若有錯誤連結存在
                logger.error("錯誤連結：" + ', '.join([link for link, status in error_links]))
                for link in internal_links:  # 將錯誤連結從內部連結中移除
                    for err_link, status in error_links:
                        if link == err_link:
                            internal_links.remove(link)
            if error_links or no_alt_links:
                record = {  # 建立錯誤連結記錄
                    'depth': current_depth,
                    'url': url,
                    'error_links': error_links,
                    'no_alt_links': no_alt_links,
                }
                all_err_links.append(record)  # 將錯誤連結記錄加入全體錯誤連結記錄中

            # 將正確的內部連結加入待檢查的連結
            for link in internal_links:
                if (
                    link and link.startswith((start_url, '/')) and current_depth <= depth_limit
                ):  # 若連結為內部連結且深度未達到指定深度
                    absolute_link = urljoin(url, link).strip()  # 取得絕對連結, 並去除前後空白
                    queue.append((absolute_link, current_depth + 1))  # 將絕對連結加入待檢查的連結
    return all_err_links


def report(report_folder, filename, result) -> None:
    '''產生xlsx報告'''
    xlsx_name = os.path.join(report_folder, f"{filename}.xlsx")  # 報告檔路徑
    workbook = openpyxl.Workbook()  # 利用 Workbook 建立一個新的工作簿
    sheet = workbook.worksheets[0]  # 取得第一個工作表
    field_names = ('層數', '網頁', '錯誤連結', '狀態碼或錯誤訊息')  # 欄位名稱
    sheet.append(field_names)  # 寫入欄位名稱
    for column in range(1, sheet.max_column + 1):  # 設定第一列的第1欄到最後1欄
        sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, column).font = Font(bold=True)  # 設定粗體
    sheet.row_dimensions[1].height = 20  # 設定第一列高度
    for rec in result:
        for link, status in rec['error_links']:
            sheet.append([rec['depth'], rec['url'], link, status])  # 寫入一列 4 個欄位
        for link in rec['no_alt_links']:
            sheet.append([rec['depth'], rec['url'], link, "圖片沒有 alt 屬性"])
    for col in ['B', 'C', 'D']:
        sheet.column_dimensions[col].width = 70  # 設定 B、C、D 欄寬度
    for row in range(1, sheet.max_row + 1):  # 設定第一列到最後一列
        sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')  # 第1欄設定置中
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).style = "Hyperlink"  # 設定超連結樣式
        sheet.cell(row, 2).hyperlink = sheet.cell(row, 2).value
        sheet.cell(row, 2).alignment = Alignment(vertical='center', wrap_text=True)  # 第2欄設定自動換行

        sheet.cell(row, 3).style = "Hyperlink"
        if not sheet.cell(row, 3).value.startswith(('http', '//')):
            sheet.cell(row, 3).hyperlink = urljoin(sheet.cell(row, 2).value, sheet.cell(row, 3).value)
        else:
            sheet.cell(row, 3).hyperlink = sheet.cell(row, 3).value
        sheet.cell(row, 3).alignment = Alignment(vertical='center', wrap_text=True)  # 第3欄設定自動換行

        sheet.cell(row, 4).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True
        )  # 第4欄設定自動換行

        for column in range(1, sheet.max_column + 1):
            sheet.cell(row, column).fill = (
                PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
                if row % 2 == 0
                else PatternFill(fill_type=None)
            )  # 設定偶數列的背景顏色
    sheet.freeze_panes = sheet['A2']  # 設定凍結第一列
    workbook.save(xlsx_name)  # 儲存檔案


def create_logger(log_folder, filename) -> logging.Logger:
    """建立 logger"""
    logger = logging.getLogger(filename)  # 取得 logger
    logger.setLevel(logging.INFO)  # 設定日誌記錄等級(共有：DEBUG,INFO,WARNING,ERROR,CRITICAL)
    formatter = logging.Formatter(
        '[%(asctime)s - %(levelname)s] %(message)s',
        datefmt='%Y%m%d %H:%M:%S',
    )  # 設定 log 格式為：[日期時間 - 等級] 日誌訊息，其中的日期時間格式為 年月日 時分秒、等級為 INFO...5種

    # 設定終端機寫入
    console_handle = logging.StreamHandler()
    console_handle.setLevel(logging.INFO)
    console_handle.setFormatter(formatter)
    # 設定檔案寫入
    logname = os.path.join(log_folder, f"{filename}.log")  # log 檔路徑
    file_handle = logging.FileHandler(logname, 'w', 'utf-8')
    file_handle.setLevel(logging.INFO)
    file_handle.setFormatter(formatter)
    # 為 log 綁定終端機與檔案
    logger.addHandler(console_handle)
    logger.addHandler(file_handle)
    return logger


def read_config(cfg_file) -> dict:
    # 讀取設定檔 config.yaml
    yaml = YAML()
    yaml.preserve_quotes = True
    yaml.indent(mapping=2, sequence=4, offset=2)
    with open(cfg_file, 'r', encoding='UTF-8') as f:
        setting = yaml.load(f)
    return setting


def create_config(cfg_file) -> dict:
    # 建立設定檔 config.yaml
    yaml = YAML()
    yaml.preserve_quotes = True
    yaml.indent(mapping=2, sequence=4, offset=2)
    with open(cfg_file, 'w', encoding='UTF-8') as f:
        setting = dict()
        setting['layer'] = 3  # 定義連結檢查層數
        setting['timeout'] = 8  # 定義逾時秒數
        setting['alt_must'] = 'no'  # 定義是否必須偵測圖片的 alt 屬性
        setting['headers'] = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6,zh-CN;q=0.5,la;q=0.4',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Cookie': '_ga_SNR7NPLEYG=GS1.1.1651249603.1.0.1651250646.0; _ga_BGEHGPV3SB=GS1.1.1707539001.1.1.1707539323.0.0.0; _gid=GA1.3.1607128698.1707801742; _ga=GA1.1.381372473.1651249604; _ga_Q0EL30K2K5=GS1.1.1707801741.1.0.1707801770.0.0.0; _ga_54MVLT2EZN=GS1.1.1707843944.5.1.1707843969.0.0.0; __RequestVerificationToken_L05ldFNlcnZpY2Vz0=MrnqY4BqFXwyAR3uGWq5prQZPEwGWyzIJgIpuGFLyP8hqJ6eLKM9EWlC8NVA4MZqmyjtxmWT-9ZtzrO04NCTXcM_njimY7J0_WFWHlyWtzE1',
            'Sec-Ch-Ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        }
        setting['avoid_urls'] = [
            "https://accessibility.moda.gov.tw/Applications/DetectLog/157716",
            "https://accessibility.moda.gov.tw/Home/Info/",
            "https://accessibility.moda.gov.tw/Download/Detail/1375?Category=52",
            "https://www.ndc.gov.tw/cp.aspx?n=32A75A78342B669D]",
        ]
        setting['scan_urls'] = ["https://www.ncut.edu.tw/"]  # 定義想要檢查的網址
        yaml.dump(setting, f)
        return setting


# 讀取設定檔 config.yaml，若設定檔存在則讀取，否則建立設定檔
config_file = 'config.yaml'
if os.path.exists(config_file):
    setting = read_config(config_file)
else:
    print("首次執行，為您建立所需環境，請稍待...")
    setting = create_config(config_file)
    print("要改變掃描網址請變更 config.yaml 檔中 scan_urls 項目下的網址。")
    input("請您再次執行程式，謝謝！按任意鍵離開...")
    exit()
if not setting.get('rpt_folder'):
    setting['rpt_folder'] = os.path.join(os.environ['USERPROFILE'], 'Documents')  # 預設檔案存放目錄為【我的文件】
# 設定全域變數
LAYER = setting.get('layer')  # 定義連結檢查層數
TIMEOUT = setting.get('timeout')  # 定義逾時秒數
ALT_MUST = setting.get('alt_must')  # 定義是否必須偵測圖片的 alt 屬性
HEADERS = setting.get('headers')  # 定義連線標頭
AVOID_URLS = setting.get('avoid_urls')  # 定義避免檢查的網址
SCAN_URLS = setting.get('scan_urls')  # 定義想要檢查的網址
DOC_FOLDER = setting.get('rpt_folder')  # 定義報告存放目錄
# 其它設定
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # 關閉 SSL 不安全的警告訊息
logging.getLogger('requests').setLevel(logging.ERROR)  # 設定 requests 的 log 等級
logging.captureWarnings(True)  # 捕捉 py waring message
visited_link = dict()  # 存放已經檢查過的連結與狀態碼

# ---------------- 主程式 ----------------
start_time = datetime.now()  # 取得開始時間
for start_url in SCAN_URLS:
    url_domain = urlparse(start_url).netloc  # 取得網域
    current_time = datetime.now()  # 取得目前時間
    filename = f"{url_domain}_{current_time.strftime('%m%d_%H%M')}"  # 建立檔案名稱，格式為 domain_月日_時分

    logger = create_logger(DOC_FOLDER, filename)  # 建立 logger
    logger.info(f"=》開始掃描 {url_domain}...")
    result = queued_link_check(start_url, LAYER)  # 根據網址進行指定層數的連結檢查
    if result:
        report(DOC_FOLDER, filename, result)  # 產生報告
        pprint(result)  # 顯示錯誤連結
        err_count = sum([len(rec['error_links']) for rec in result])
        logger.info(f"=》掃描 {url_domain} 完成，共有 {err_count} 個錯誤連結。")
        logger.info(f"=》報告存放於 {DOC_FOLDER}...")
    else:
        logger.info(f"=》太棒了！{url_domain} 沒有錯誤連結。")
end_time = datetime.now()  # 取得結束時間
hours, remainder = divmod((end_time - start_time).seconds, 3600)
minutes, seconds = divmod(remainder, 60)
print(f"=》總共花費時間：{hours} 小時 {minutes} 分鐘 {seconds} 秒。")

# 移除所有 logger 的 handlers
for handler in logger.handlers[:]:
    handler.close()
    logger.removeHandler(handler)

os.system(f'explorer {DOC_FOLDER}')  # 開啟檔案存放目錄
