import datetime
import re
from collections import deque
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font

LAYER = 4  # 定義檢查連結的層數
TIMEOUT = 8  # 定義連線逾時秒數
HEADERS = {
    'content-type': 'text/html; charset=utf-8',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
}  # 定義連線標頭


def is_valid(url, link) -> bool:
    '''檢查連結是否有效'''
    fqdn = urljoin(url, link)
    try:
        print(" " * 140, end="\r")  # 清除目前列
        print(f"檢查連結: {link} ... ", flush=True, end="")  # 顯示目前檢查的連結，不快取強制輸出後不換行
        response = requests.get(
            fqdn, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True
        )  # 發送 GET 請求，其實發 HEAD 請求會比較好，但會返回 403 很奇怪，可能是 Server 端設定錯誤
        status = response.status_code  # 取得狀態碼
        print(f"狀態: {status}", flush=True, end="\r")  # 顯示狀態碼，游標至目前列首
        return status in (200, 301, 302)  # 若狀態碼為 200, 301, 302 則回傳 True
    except requests.exceptions.Timeout as e:
        status = f"連線逾時：{link}  錯誤訊息：{e}"
    except requests.exceptions.TooManyRedirects as e:
        status = f"重新導向次數過多：{link}  錯誤訊息：{e}"
    except requests.exceptions.RequestException as e:
        status = f"無法取得此網頁內容：{link}  錯誤訊息：{e}"
    except requests.exceptions.ConnectionError as e:
        status = f"無法連線至此網頁：{link}  錯誤訊息：{e}"
    except requests.exceptions.HTTPError as e:
        status = f"HTTP 錯誤：{link}  錯誤訊息：{e}"
    except Exception as e:
        status = f"其它錯誤：{link}  錯誤訊息：{e}"

    print(f"狀態: {status}", flush=True, end="\r")  # 顯示狀態碼，游標至目前列首
    return False  # 若發生錯誤，則回傳 False


def get_links(url) -> tuple:
    '''取得網頁中的所有內部連結與外部連結'''
    try:
        response = requests.get(url, headers=HEADERS, timeout=TIMEOUT)  # 發送 GET 請求
        soup = BeautifulSoup(response.content, 'lxml')  # 使用 lxml 解析器解析 HTML 內容
        all_links = []  # 取得所有的連結
        inter_links = []  # 內部連結
        outer_links = []  # 外部連結

        hrefs = [tag.get('href') for tag in soup.find_all(href=True)]  # 找到所有具有 href 屬性的屬性值
        srcs = [tag.get('src') for tag in soup.find_all(src=True)]  # 找到所有具有 src 屬性的屬性值
        all_links.extend(hrefs)  # 合併 href 屬性值到所有連結
        all_links.extend(srcs)  # 合併 src 屬性值到所有連結
        all_links = list(set(all_links))  # 去除重複的連結
        for link in all_links:
            if link.startswith(url) or link.startswith("/"):  # 若連結為內部連結
                inter_links.append(link)
            elif link.startswith(('http://', 'https://')):  # 若連結為外部連結
                outer_links.append(link)
        return (inter_links, outer_links)  # 回傳內部連結與外部連結
    except requests.RequestException:
        return ([], [])  # 若發生錯誤，則回傳空串列


def queued_link_check(start_url, depth_limit=1) -> list:
    '''使用雙向佇列結構儲存網站中的連結'''
    visited = set()  # 存放已經檢查過的連結
    queue = deque([(start_url, 0)])  # 存放待檢查的連結，設定目前深度為 0
    all_err_links = []  # 存放所有錯誤連結

    while queue:
        url, current_depth = queue.popleft()  # 取得待檢查的連結
        match = re.search(r".*\.\w{2,4}$", url)  # 是否符合具有副檔名的.xxx或.xxxx (例如: .7z, .php, .aspx)
        if match and not url.endswith(('html', 'htm')):  # 若連結為檔案且非網頁檔
            continue  # 跳過
        if url not in visited and current_depth <= depth_limit:  # 若連結未檢查過且深度未達到指定深度
            err_links = []  # 存放此次 url 的錯誤連結
            print(f"\n第 {current_depth} 層連結： {url}")
            visited.add(url)  # 將連結加入已檢查的集合
            inter_links, outer_links = get_links(url)  # 取得內部連結與外部連結
            err_inter_links = [link1 for link1 in inter_links if not is_valid(url, link1)]  # 取得錯誤的內部連結
            err_outer_links = [link2 for link2 in outer_links if not is_valid(url, link2)]  # 取得錯誤的外部連結
            err_links.extend(err_inter_links) if err_inter_links else err_links  # 將錯誤的內部連結加入錯誤連結集合
            err_links.extend(err_outer_links) if err_outer_links else err_links  # 將錯誤的外部連結加入錯誤連結集合
            err_links = list(set(err_links))  # 去除重複的錯誤連結
            exist_inter_links = [x for x in inter_links if x not in err_inter_links]  # 取得正確的內部連結
            # exist_outer_links = [x for x in outer_links if x not in err_outer_links]
            if err_links:
                rec = {
                    'depth': current_depth,
                    'url': url,
                    'err_links': err_links,
                }
                all_err_links.append(rec)  # 將錯誤連結加入錯誤連結集合

            # 將正確的內部連結加入待檢查的連結
            for link in exist_inter_links:
                if (
                    link and link.startswith((start_url, '/')) and current_depth <= depth_limit
                ):  # 若連結為內部連結且深度未達到指定深度
                    absolute_link = urljoin(url, link)  # 取得絕對連結
                    queue.append((absolute_link, current_depth + 1))  # 將絕對連結加入待檢查的連結
    return all_err_links


def report(start_url, result):
    url_domain = urlparse(start_url).netloc  # 取得網域
    current_time = datetime.datetime.now()  # 取得目前時間
    # 建立檔案名稱，格式為 domain_月日_時分
    filename = f"{url_domain}_{current_time.strftime('%m%d_%H%M')}"

    xlsxname = f"{filename}.xlsx"
    workbook = openpyxl.Workbook()  # 利用 Workbook 建立一個新的工作簿
    sheet = workbook.worksheets[0]  # 取得第一個工作表
    fieldnames = ('層數', '網頁', '錯誤連結')  # 欄位名稱
    sheet.append(fieldnames)  # 寫入欄位名稱
    for column in range(1, sheet.max_column + 1):  # 設定第一欄到最後一欄
        sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, column).font = Font(bold=True)  # 設定粗體
    sheet.row_dimensions[1].height = 20  # 設定第一列高度
    for rec in result:
        for link in rec['err_links']:
            sheet.append([rec['depth'], rec['url'], link])  # 寫入一列3個欄位
    sheet.column_dimensions['B'].width = 80  # 設定B欄寬度
    sheet.column_dimensions['C'].width = 80  # 設定C欄寬度
    for row in range(1, sheet.max_row + 1):  # 設定第一列到最後一列
        sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')  # 第一欄設定置中
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).style = "Hyperlink"  # 設定超連結樣式
        sheet.cell(row, 2).hyperlink = sheet.cell(row, 2).value  # 設定超連結
        sheet.cell(row, 2).alignment = Alignment(vertical='center', wrap_text=True)  # 設定自動換行
        sheet.cell(row, 3).style = "Hyperlink"
        sheet.cell(row, 3).hyperlink = urljoin(sheet.cell(row, 2).value, sheet.cell(row, 3).value)
        sheet.cell(row, 3).alignment = Alignment(vertical='center', wrap_text=True)
    sheet.freeze_panes = 'A2'  # 設定凍結第一列
    workbook.save(xlsxname)  # 儲存檔案


# 主程式
if __name__ == "__main__":
    default_url = "https://cc.ncut.edu.tw/"
    # default_url = "https://lgc.ncut.edu.tw/"
    start_url = input(f"請輸入網址({default_url})：")
    start_url = default_url if start_url == "" else start_url  # 若未輸入網址，則使用預設網址

    result = queued_link_check(start_url, LAYER)  # 根據網址進行指定層數的連結檢查
    report(start_url, result)  # 產生報告
