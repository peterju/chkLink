import datetime
import logging
import os
import re
from collections import deque
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill

# from openpyxl.worksheet.hyperlink import Hyperlink
# https://titangene.github.io/article/python-logging.html

# ---------------- 使用者設定 ----------------
WANTED_URL = ("https://www.ncut.edu.tw/", "https://cc.ncut.edu.tw/", "https://lgc.ncut.edu.tw/")  # 要檢查的網址
# WANTED_URL = ("https://www.ncut.edu.tw/",)  # 若檢查的網址只有一個，則需在最後加上逗號，否則會被當成字串處理
LAYER = 3  # 定義檢查連結的層數
TIMEOUT = 8  # 定義連線逾時秒數
ALT_MUST = False  # 圖片是否必須有 alt 屬性
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
}  # 定義連線標頭

# ---------------- 系統設定 ----------------
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # 關閉 SSL 不安全的警告訊息
doc_folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Documents')  # 預設檔案存放目錄為【我的文件】


def is_valid(url, link) -> bool:
    '''檢查連結是否有效'''
    fqdn = urljoin(url, link).strip()  # 取得絕對連結, 並去除前後空白
    protocol = urlparse(fqdn).scheme  # 取得協定

    try:
        print(f"檢查: {link} ... ", flush=True, end="\r")  # 顯示目前檢查的連結，不快取強制輸出後不換行
        response = requests.get(
            fqdn, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, stream=True, verify=False
        )  # 發送 GET 請求，其實發 HEAD 請求會比較好，但會返回 403 很奇怪，可能是 Server 端設定錯誤
        status = str(response.status_code)  # 取得狀態碼
    except requests.exceptions.Timeout as e:
        status = f"連線逾時：{link}  錯誤訊息：{e}"
    except requests.exceptions.TooManyRedirects as e:
        status = f"重新導向次數過多：{link}  錯誤訊息：{e}"
    except requests.exceptions.RequestException as e:
        status = f"無法取得此網頁內容：{link}  錯誤訊息：{e}"
    except requests.exceptions.ConnectionError as e:
        status = f"無法連線至此網頁：{link}  錯誤訊息：{e}"
    except requests.exceptions.HTTPError as e:
        status = f"HTTP 錯誤：{link}  錯誤訊息：{e}"
    except Exception as e:
        status = f"其它錯誤：{link}  錯誤訊息：{e}"
    if status == "403":
        status = f"請求被拒絕：{status}"
    elif status == "404":
        status = f"請求失敗：{status}"
    elif protocol == "http":
        status = f"使用 http 協定並不安全，{status}"
    if "200" in status:
        logger.info(f"檢查: {link} ... 狀態: {status}")
    else:
        logger.error(f"檢查: {link} ... 狀態: {status}")
    return status


def get_links(url) -> tuple:
    '''取得網頁中的所有內部連結與外部連結'''
    try:
        response = requests.get(url.strip(), headers=HEADERS, timeout=TIMEOUT, verify=False)  # 發送 GET 請求
        soup = BeautifulSoup(response.content, 'lxml')  # 使用 lxml 解析器解析 HTML 內容
        all_links = []  # 取得所有的連結
        inter_links = []  # 內部連結
        outer_links = []  # 外部連結
        no_alt_links = []  # 圖片沒有 alt 屬性的連結

        hrefs = [tag.get('href').strip() for tag in soup.find_all(href=True)]  # 找到所有具有 href 屬性的屬性值
        srcs = [tag.get('src').strip() for tag in soup.find_all(src=True)]  # 找到所有具有 src 屬性的屬性值
        if ALT_MUST:  # 若須偵測圖片的 alt 屬性
            no_alt_links = [
                tag.get('src') for tag in soup.find_all('img', alt=False)
            ]  # 找到所有圖片沒有 alt 屬性的標籤
        all_links.extend(hrefs)  # 合併 href 屬性值到所有連結
        all_links.extend(srcs)  # 合併 src 屬性值到所有連結
        all_links = list(set(all_links))  # 去除重複的連結
        for link in all_links:
            # if link.startswith(url) or link.startswith("/"):  # 若連結為內部連結
            if link.startswith((url, '/')):  # 若連結為內部連結
                inter_links.append(link)
            elif link.startswith(('http://', 'https://')):  # 若連結為外部連結
                outer_links.append(link)
        return (inter_links, outer_links, no_alt_links)  # 回傳內部連結、外部連結、圖片沒有 alt 屬性的連結
    except requests.RequestException:
        return ([], [], [])  # 若發生錯誤，則回傳空串列


def queued_link_check(start_url, depth_limit=1) -> list:
    '''使用雙向佇列結構儲存網站中的連結'''
    visited = set()  # 存放已經檢查過的連結
    queue = deque([(start_url, 0)])  # 存放待檢查的連結，設定目前深度為 0
    all_err_links = []  # 存放所有錯誤連結

    while queue:
        url, current_depth = queue.popleft()  # 取得待檢查的連結
        match = re.search(r".*\.\w{2,4}$", url)  # 是否符合具有副檔名的.xxx或.xxxx (例如: .7z, .php, .aspx)
        if match and not url.endswith(('html', 'htm')):  # 若連結為檔案且非網頁檔
            continue  # 跳過
        if url not in visited and current_depth <= depth_limit:  # 若連結未檢查過且深度未達到指定深度
            err_links = []  # 存放此次 url 的錯誤連結
            logger.info(f"第 {current_depth} 層連結： {url}")
            visited.add(url)  # 將連結加入已檢查的集合
            inter_links, outer_links, no_alt_links = get_links(url)  # 取得內部連結、外部連結與沒有alt的連結
            err_inter_links = []  # 存放錯誤的內部連結
            for link in inter_links:
                status = is_valid(url, link)
                if '200' not in status:
                    err_inter_links.append((link, status))  # 取得錯誤的內部連結
            err_outer_links = []  # 存放錯誤的外部連結
            for link in outer_links:
                status = is_valid(url, link)
                if '200' not in status:
                    err_outer_links.append((link, status))  # 取得錯誤的外部連結

            err_links.extend(err_inter_links) if err_inter_links else err_links  # 將錯誤的內部連結加入錯誤連結集合
            err_links.extend(err_outer_links) if err_outer_links else err_links  # 將錯誤的外部連結加入錯誤連結集合
            err_links = list(set(err_links))  # 去除重複的錯誤連結

            if err_links:  # 若有錯誤連結存在
                for link in inter_links:  # 將錯誤連結從內部連結中移除
                    for err_link, status in err_links:
                        if link == err_link:
                            inter_links.remove(link)

            rec = {  # 建立錯誤連結記錄
                'depth': current_depth,
                'url': url,
                'err_links': err_links,
                'no_alt_links': no_alt_links,
            }
            all_err_links.append(rec)  # 將錯誤連結記錄加入全體錯誤連結記錄中

            # 將正確的內部連結加入待檢查的連結
            for link in inter_links:
                if (
                    link and link.startswith((start_url, '/')) and current_depth <= depth_limit
                ):  # 若連結為內部連結且深度未達到指定深度
                    absolute_link = urljoin(url, link).strip()  # 取得絕對連結, 並去除前後空白
                    queue.append((absolute_link, current_depth + 1))  # 將絕對連結加入待檢查的連結
    return all_err_links


def report(rpt_folder, filename, result) -> None:
    '''產生xlsx報告'''
    xlsxname = os.path.join(rpt_folder, f"{filename}.xlsx")  # 報告檔路徑
    workbook = openpyxl.Workbook()  # 利用 Workbook 建立一個新的工作簿
    sheet = workbook.worksheets[0]  # 取得第一個工作表
    fieldnames = ('層數', '網頁', '錯誤連結', '狀態碼或錯誤訊息')  # 欄位名稱
    sheet.append(fieldnames)  # 寫入欄位名稱
    for column in range(1, sheet.max_column + 1):  # 設定第一列的第1欄到最後1欄
        sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, column).font = Font(bold=True)  # 設定粗體
    sheet.row_dimensions[1].height = 20  # 設定第一列高度
    for rec in result:
        for link, status in rec['err_links']:
            sheet.append([rec['depth'], rec['url'], link, status])  # 寫入一列 4 個欄位
        for link in rec['no_alt_links']:
            sheet.append([rec['depth'], rec['url'], link, "圖片沒有 alt 屬性"])
    sheet.column_dimensions['B'].width = 70  # 設定B欄寬度
    sheet.column_dimensions['C'].width = 70  # 設定C欄寬度
    sheet.column_dimensions['D'].width = 70  # 設定D欄寬度
    for row in range(1, sheet.max_row + 1):  # 設定第一列到最後一列
        sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')  # 第1欄設定置中
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).style = "Hyperlink"  # 設定超連結樣式
        sheet.cell(row, 2).hyperlink = sheet.cell(row, 2).value
        sheet.cell(row, 2).alignment = Alignment(vertical='center', wrap_text=True)  # 第2欄設定自動換行

        sheet.cell(row, 3).style = "Hyperlink"
        if sheet.cell(row, 3).value.startswith('/'):
            sheet.cell(row, 3).hyperlink = urljoin(sheet.cell(row, 2).value, sheet.cell(row, 3).value)
        else:
            sheet.cell(row, 3).hyperlink = sheet.cell(row, 3).value
        sheet.cell(row, 3).alignment = Alignment(vertical='center', wrap_text=True)  # 第3欄設定自動換行

        sheet.cell(row, 4).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True
        )  # 第4欄設定自動換行

        for column in range(1, sheet.max_column + 1):
            sheet.cell(row, column).fill = (
                PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
                if row % 2 == 0
                else PatternFill(fill_type=None)
            )  # 設定偶數列的背景顏色
    sheet.freeze_panes = sheet['A2']  # 設定凍結第一列
    workbook.save(xlsxname)  # 儲存檔案


def create_logger(log_folder, filename) -> logging.Logger:
    """建立 logger"""
    logname = os.path.join(log_folder, f"{filename}.log")  # log 檔路徑
    logging.captureWarnings(True)  # 捕捉 py waring message
    logger = logging.getLogger()
    # 若要替換現有的logger範例
    # logger = logging.getLogger()      # root logger
    # for hdlr in logger.handlers[:]:   # remove all old handlers
    #     logger.removeHandler(hdlr)
    # logger.addHandler(fileh)          # set the new handler
    logger.setLevel(logging.INFO)  # 共有： DEBUG,INFO,WARNING,ERROR,CRITICAL (預設成 INFO)
    formatter = logging.Formatter(
        # '[%(levelname)1.1s %(asctime)s %(module)s:%(lineno)d] %(message)s', datefmt='%Y%m%d %H:%M:%S'
        '[%(asctime)s - %(levelname)s] %(message)s',
        datefmt='%Y%m%d %H:%M:%S',
    )
    # logging.basicConfig(
    #     level=logging.INFO,
    #     filename='history.log',
    #     filemode='w',
    #     format='%(asctime)s - %(levelname)s : %(message)s',
    #     datefmt='%Y%m%d %H:%M:%S',
    # )
    # 設定終端機寫入
    console_handle = logging.StreamHandler()
    console_handle.setLevel(logging.INFO)
    console_handle.setFormatter(formatter)
    # 設定檔案寫入
    file_handle = logging.FileHandler(logname, 'w', 'utf-8')
    file_handle.setLevel(logging.INFO)
    file_handle.setFormatter(formatter)
    # 為 log 綁定終端機與檔案
    logger.addHandler(console_handle)
    logger.addHandler(file_handle)
    return logger


# ---------------- 主程式 ----------------
# default_url = "https://www.ncut.edu.tw/"
# start_url = input(f"請輸入網址({default_url})：")
# start_url = default_url if start_url == "" else start_url  # 若未輸入網址，則使用預設網址

for start_url in WANTED_URL:
    url_domain = urlparse(start_url).netloc  # 取得網域
    current_time = datetime.datetime.now()  # 取得目前時間
    # 建立檔案名稱，格式為 domain_月日_時分
    filename = f"{url_domain}_{current_time.strftime('%m%d_%H%M')}"

    logging.getLogger('requests').setLevel(logging.ERROR)  # 設定 requests 的 log 等級
    logger = create_logger(doc_folder, filename)  # 指定 log 存放的目錄
    logger.info(f"=》開始掃描 {url_domain}...")
    result = queued_link_check(start_url, LAYER)  # 根據網址進行指定層數的連結檢查
    if result:
        report(doc_folder, filename, result)  # 產生報告
    logger.info(f"=》掃描完成 {url_domain}...")
    logger.info(f"=》報告已存放於 {doc_folder}...")
os.system(f'explorer {doc_folder}')
