import logging
import os
import re
import subprocess
import threading
import tkinter as tk
from collections import deque
from datetime import datetime
from idlelib.tooltip import Hovertip
from tkinter import filedialog, messagebox, scrolledtext
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
import ttkbootstrap as ttk  # https://github.com/israel-dryer/ttkbootstrap
import urllib3
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from ruamel.yaml import YAML

# from pprint import pprint
# from openpyxl.worksheet.hyperlink import Hyperlink
# https://titangene.github.io/article/python-logging.html


def is_valid_link(base_url, link) -> bool:
    '''檢查連結是否有效'''
    global logger
    global visited_link
    full_url = urljoin(base_url, link).strip()  # 取得絕對連結, 並去除前後空白
    if full_url in AVOID_URLS:  # 若連結為避免檢查的網址
        status = "200 (避免檢查的網址)"
    elif full_url in visited_link:  # 若連結已經檢查過
        status = visited_link[full_url]  # 取得狀態碼
    else:
        protocol = urlparse(full_url).scheme  # 取得協定
        try:
            # print(f"檢查: {link} ... ", flush=True, end="\r")  # 顯示目前檢查的連結，不快取強制輸出後不換行
            if protocol == "https":
                response = requests.get(
                    full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, stream=True
                )  # 發送 GET 請求，其實發 HEAD 請求會比較好，但會返回 403 很奇怪，可能是 Server 端設定錯誤
            else:
                response = requests.get(
                    full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, stream=True, verify=False
                )
            status = str(response.status_code)  # 取得狀態碼
            real_url = response.url  # 取得網頁的實際連結, 避免重定向造成誤判
            domain1 = urlparse(full_url).netloc  # 取得原始連結的網域
            domain2 = urlparse(real_url).netloc  # 取得實際連結的網域
            # domain1 = '.'.join(urlparse(full_url).netloc.split('.')[1:])  # 取得原始連結的父網域
            # domain2 = '.'.join(urlparse(real_url).netloc.split('.')[1:])  # 取得實際連結的父網域
            if status == '200' and domain1 != domain2:  # 若原始連結與實際連結的網域不同(例如: 302 暫時定向到別的網域)
                if protocol == "https":
                    response = requests.get(
                        full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=False, stream=True
                    )
                else:
                    response = requests.get(
                        full_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=False, stream=True, verify=False
                    )
                status = f'200 重定向：{str(response.status_code)} 到 {real_url}'  # 設定狀態
        except requests.exceptions.Timeout as e:
            status = f"連線逾時：{link}  錯誤訊息：{e}"
        except requests.exceptions.TooManyRedirects as e:
            status = f"重新導向次數過多：{link}  錯誤訊息：{e}"
        except requests.exceptions.SSLError as e:
            status = f"SSL 錯誤或憑證不正確：{link}  錯誤訊息：{e}"
        except requests.exceptions.RequestException as e:
            status = f"無法取得此網頁內容：{link}  錯誤訊息：{e}"
        except requests.exceptions.ConnectionError as e:
            status = f"無法連線至此網頁：{link}  錯誤訊息：{e}"
        except requests.exceptions.HTTPError as e:
            status = f"HTTP 錯誤：{link}  錯誤訊息：{e}"
        except Exception as e:
            status = f"其它錯誤：{link}  錯誤訊息：{e}"
        if status == "403":
            status = f"{status} 請求被拒絕"
        elif status == "404":
            status = f"{status} 請求失敗"
        elif protocol == "http":
            status = f"{status} 但使用 http 協定並不安全"
        visited_link[full_url] = status  # 將連結與狀態碼加入已檢查的集合

    msg = f"檢查: {link} ... 狀態: {status}"
    if "200" in status:
        logger.info(msg)
        log_console.insert(tk.END, msg + "\n", "INFO")
    else:
        logger.error(msg)
        log_console.insert(tk.END, msg + "\n", "ERROR")
    log_console.see(tk.END)  # 捲動 log_console 至最後一行
    return status


def get_links(domain, url) -> tuple:
    '''取得網頁中的所有內部連結與外部連結'''
    global logger
    try:
        protocol = urlparse(url).scheme  # 取得協定
        if protocol == "https":
            response = requests.get(
                url.strip(), headers=HEADERS, timeout=TIMEOUT, allow_redirects=True
            )  # 發送 GET 請求
        else:
            response = requests.get(url.strip(), headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, verify=False)
        real_url = response.url  # 取得網頁的實際連結, 避免重定向造成誤判
        real_domain = urlparse(real_url).netloc  # 取得網域
        if domain != real_domain:  # 若原始連結與實際連結的網域不同
            return ([], [], [])  # 回傳空串列
        soup = BeautifulSoup(response.content, 'html.parser')  # 使用 BeautifulSoup 解析網頁內容
        all_links = []  # 取得所有的連結
        internal_links, external_links, no_alt_links = [], [], []  # 初始化內部連結、外部連結與沒有 alt 的連結

        hrefs = [tag.get('href').strip() for tag in soup.find_all(href=True)]  # 找到所有具有 href 屬性的屬性值
        srcs = [tag.get('src').strip() for tag in soup.find_all(src=True)]  # 找到所有具有 src 屬性的屬性值
        if ALT_MUST:  # 若須偵測圖片的 alt 屬性
            no_alt_links = [
                tag.get('src') for tag in soup.find_all('img', alt=False)
            ]  # 找到所有圖片沒有 alt 屬性的標籤
        all_links.extend(hrefs)  # 合併 href 屬性值到所有連結
        all_links.extend(srcs)  # 合併 src 屬性值到所有連結
        all_links = list(set(all_links))  # 去除重複的連結
        for link in all_links:
            link_domain = urlparse(link).netloc  # 取得連結的網域
            if link.startswith(('#', 'javascript', 'mailto')):  # 若連結為錨點、javascript 或郵件連結，則跳過
                continue
            elif link_domain == domain and link.startswith('/'):  # 若連結為 目前網址與 / 開頭的為內部連結
                internal_links.append(link)
            elif link.startswith(('http', '//')):  # 若連結為 http 或 // 開頭的為外部連結
                external_links.append(link)
            else:  # 其它的為內部連結，例如檔名開頭的
                internal_links.append(link)
        return (internal_links, external_links, no_alt_links)  # 回傳內部連結、外部連結、圖片沒有 alt 屬性的連結
    # except requests.RequestException as e:
    except Exception as e:
        msg = f"無法取得此網頁內容：{url}  錯誤訊息：{e}"
        logger.error(msg)
        log_console.insert(tk.END, msg + "\n", "ERROR")
        log_console.see(tk.END)  # 捲動 log_console 至最後一行
        return ([], [], [])  # 若發生錯誤，則回傳空串列


def analysis_func(start_urls, depth_limit=1):
    log_console.delete(1.0, tk.END)  # 清空 log_console
    if len(start_urls) > 0:
        analysis_btn.config
        threading.Thread(
            target=queued_link_check, args=(start_urls, depth_limit)
        ).start()  # 使用 threading 在背景執行 queued_link_check 函式
    else:
        messagebox.showwarning("資訊", "請輸入網址才能進行失效連結掃描！")


def queued_link_check(start_urls, depth_limit=1) -> list:
    '''使用雙向佇列結構儲存網站中的連結'''
    global logger
    start_time = datetime.now()  # 取得開始時間
    for start_url in start_urls.split(','):  # 以逗號分隔多個網址
        start_url = start_url.strip()  # 去除前後空白
        url_domain = urlparse(start_url).netloc  # 取得網域
        current_time = datetime.now()  # 取得目前時間
        filename = f"{url_domain}_{current_time.strftime('%m%d_%H%M')}"  # 建立檔案名稱，格式為 domain_月日_時分
        logger = create_logger(DOC_FOLDER, filename)  # 建立 logger
        msg = f"=》開始掃描 {url_domain}..."
        logger.info(msg)
        log_console.insert(tk.END, msg + "\n", "INFO")
        log_console.see(tk.END)  # 捲動 log_console 至最後一行

        visited_url = set()  # 存放已經檢查過的連結
        queue = deque([(start_url, 0)])  # 存放待檢查的連結，設定目前深度為 0
        all_err_links = []  # 存放所有錯誤連結
        while queue:
            url, current_depth = queue.popleft()  # 取得待檢查的連結
            if url in AVOID_URLS:  # 若連結為避免檢查的網址
                continue  # 跳過
            match = re.search(r".*\.\w{2,4}$", url)  # 是否符合具有副檔名的.xxx或.xxxx (例如: .7z, .php, .aspx)
            if match and not url.endswith(
                ('html', 'htm', 'php', 'asp', 'aspx', 'jsp', 'tw', 'com')
            ):  # 若連結為檔案且非網頁檔
                continue  # 跳過

            if url not in visited_url and current_depth <= depth_limit:  # 若連結未檢查過且深度未達到指定深度
                internal_links, external_links, no_alt_links = [], [], []  # 初始化內部連結、外部連結與沒有 alt 的連結
                msg = f"第 {current_depth} 層連結： {url}"
                logger.info(msg)  # 顯示目前處理的連結
                log_console.insert(tk.END, msg + "\n", "INFO")
                log_console.see(tk.END)  # 捲動 log_console 至最後一行
                visited_url.add(url)  # 將連結加入已檢查的集合
                domain = urlparse(url).netloc  # 取得網域
                internal_links, external_links, no_alt_links = get_links(
                    domain, url
                )  # 取得內部連結、外部連結與沒有alt的連結
                if internal_links:
                    msg = f"內部連結：{', '.join(internal_links)}"
                    logger.info(msg)
                    log_console.insert(tk.END, msg + "\n", "INFO")
                if external_links:
                    msg = f"外部連結：{', '.join(external_links)}"
                    logger.info(msg)
                    log_console.insert(tk.END, msg + "\n", "INFO")
                if no_alt_links:
                    msg = f"沒有 alt 屬性的連結：{', '.join(no_alt_links)}"
                    logger.info(msg)
                    log_console.insert(tk.END, msg + "\n", "INFO")
                log_console.see(tk.END)  # 捲動 log_console 至最後一行
                error_internal_links = []  # 存放錯誤的內部連結
                for link in internal_links:
                    status = is_valid_link(url, link)  # 檢查連結是否有效
                    if '200' not in status:  # 若狀態碼不是 200
                        error_internal_links.append((link, status))  # 將錯誤的連結加入存放錯誤的內部連結
                error_external_links = []  # 存放錯誤的外部連結
                for link in external_links:  # 檢查外部連結是否有效
                    status = is_valid_link(url, link)  # 檢查連結是否有效
                    if '200' not in status:  # 若狀態碼不是 200
                        error_external_links.append((link, status))  # 將錯誤的連結加入存放錯誤的外部連結

                error_links = []  # 存放此次所有 url 錯誤的連結
                # 將錯誤的內部連結加入所有錯誤連結集合
                error_links.extend(error_internal_links) if error_internal_links else error_links
                # 將錯誤的外部連結加入所有錯誤連結集合
                error_links.extend(error_external_links) if error_external_links else error_links
                error_links = list(set(error_links))  # 去除重複的錯誤連結
                if error_links:  # 若有錯誤連結存在
                    msg = f"錯誤連結：{', '.join([link for link, status in error_links])}"
                    logger.error(msg)
                    log_console.insert(tk.END, msg + "\n", "ERROR")
                    log_console.see(tk.END)  # 捲動 log_console 至最後一行
                    for link in internal_links:  # 將錯誤連結從內部連結中移除
                        for err_link, status in error_links:
                            if link == err_link:
                                internal_links.remove(link)

                if error_links or no_alt_links:
                    record = {  # 建立錯誤連結記錄
                        'depth': current_depth,
                        'url': url,
                        'error_links': error_links,
                        'no_alt_links': no_alt_links,
                    }
                    all_err_links.append(record)  # 將錯誤連結記錄加入全體錯誤連結記錄中

                # 將正確的內部連結加入待檢查的連結
                for link in internal_links:
                    if (
                        link and link.startswith((start_url, '/')) and current_depth <= depth_limit
                    ):  # 若連結為內部連結且深度未達到指定深度
                        absolute_link = urljoin(url, link).strip()  # 取得絕對連結, 並去除前後空白
                        queue.append((absolute_link, current_depth + 1))  # 將絕對連結加入待檢查的連結
        # return all_err_links
        if all_err_links:
            report(DOC_FOLDER, filename, all_err_links)  # 產生報告
            err_count = sum([len(rec['error_links']) for rec in all_err_links])
            msg = f"=》掃描 {url_domain} 完成，共有 {err_count} 個錯誤連結。"
            logger.info(msg)
            log_console.insert(tk.END, msg + "\n", "SUCCESS")
            log_console.see(tk.END)  # 捲動 log_console 至最後一行

        msg = f"=》報告已存放於 {DOC_FOLDER}..."
        logger.info(msg)
        log_console.insert(tk.END, msg + "\n", "SUCCESS")
        log_console.see(tk.END)  # 捲動 log_console 至最後一行

    analysis_btn.config(state=tk.NORMAL)  # 完成後重新啟用分析按鈕
    end_time = datetime.now()  # 取得結束時間
    hours, remainder = divmod((end_time - start_time).seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    msg = f"=》全部掃描完成！共花費 {hours} 小時 {minutes} 分 {seconds} 秒。"
    log_console.insert(tk.END, msg + "\n", "SUCCESS")
    log_console.see(tk.END)  # 捲動 log_console 至最後一行

    # 移除所有 logger 的 handlers
    for handler in logger.handlers[:]:
        handler.close()
        logger.removeHandler(handler)
    del logger  # 刪除 logger

    answer = messagebox.askquestion("資訊", "掃描完成！是否要開啟檔案所在目錄？")
    if answer == 'yes':
        subprocess.Popen(f"explorer {DOC_FOLDER}")
        # subprocess.Popen(f"explorer {entry_convert_dir.get()}")


def report(report_folder, filename, result) -> None:
    '''產生xlsx報告'''
    xlsx_name = os.path.join(report_folder, f"{filename}.xlsx")  # 報告檔路徑
    workbook = openpyxl.Workbook()  # 利用 Workbook 建立一個新的工作簿
    sheet = workbook.worksheets[0]  # 取得第一個工作表
    field_names = ('層數', '網頁', '錯誤連結', '狀態碼或錯誤訊息')  # 欄位名稱
    sheet.append(field_names)  # 寫入欄位名稱
    for column in range(1, sheet.max_column + 1):  # 設定第一列的第1欄到最後1欄
        sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, column).font = Font(bold=True)  # 設定粗體
    sheet.row_dimensions[1].height = 20  # 設定第一列高度
    for rec in result:
        for link, status in rec['error_links']:
            sheet.append([rec['depth'], rec['url'], link, status])  # 寫入一列 4 個欄位
        for link in rec['no_alt_links']:
            sheet.append([rec['depth'], rec['url'], link, "圖片沒有 alt 屬性"])
    for col in ['B', 'C', 'D']:
        sheet.column_dimensions[col].width = 70  # 設定 B、C、D 欄寬度
    for row in range(1, sheet.max_row + 1):  # 設定第一列到最後一列
        sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')  # 第1欄設定置中
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).style = "Hyperlink"  # 設定超連結樣式
        sheet.cell(row, 2).hyperlink = sheet.cell(row, 2).value
        sheet.cell(row, 2).alignment = Alignment(vertical='center', wrap_text=True)  # 第2欄設定自動換行

        sheet.cell(row, 3).style = "Hyperlink"
        if not sheet.cell(row, 3).value.startswith(('http', '//')):
            sheet.cell(row, 3).hyperlink = urljoin(sheet.cell(row, 2).value, sheet.cell(row, 3).value)
        else:
            sheet.cell(row, 3).hyperlink = sheet.cell(row, 3).value
        sheet.cell(row, 3).alignment = Alignment(vertical='center', wrap_text=True)  # 第3欄設定自動換行

        sheet.cell(row, 4).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True
        )  # 第4欄設定自動換行

        for column in range(1, sheet.max_column + 1):
            sheet.cell(row, column).fill = (
                PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
                if row % 2 == 0
                else PatternFill(fill_type=None)
            )  # 設定偶數列的背景顏色
    sheet.freeze_panes = sheet['A2']  # 設定凍結第一列
    workbook.save(xlsx_name)  # 儲存檔案


def create_logger(log_folder, filename) -> logging.Logger:
    """建立 logger"""
    global logger, log_console
    logger = logging.getLogger(filename)  # 取得 logger
    logger.setLevel(logging.INFO)  # 設定日誌記錄等級(共有：DEBUG,INFO,WARNING,ERROR,CRITICAL)
    formatter = logging.Formatter(
        '[%(asctime)s - %(levelname)s] %(message)s',
        datefmt='%Y%m%d %H:%M:%S',
    )  # 設定 log 格式為：[日期時間 - 等級] 日誌訊息，其中的日期時間格式為 年月日 時分秒、等級為 INFO...5種

    # 設定終端機寫入
    # console_handle = logging.StreamHandler()
    # console_handle.setLevel(logging.INFO)
    # console_handle.setFormatter(formatter)
    # 設定檔案寫入
    logname = os.path.join(log_folder, f"{filename}.log")  # log 檔路徑
    file_handle = logging.FileHandler(logname, 'w', 'utf-8')
    file_handle.setLevel(logging.INFO)
    file_handle.setFormatter(formatter)

    # 為 log 綁定終端機與檔案
    # logger.addHandler(console_handle)
    logger.addHandler(file_handle)
    return logger


def save_config() -> None:
    '''儲存設定檔'''
    global HEADERS, TIMEOUT, LAYER, ALT_MUST, SCAN_URLS, AVOID_URLS
    HEADERS = dict()  # 定義連線標頭
    for header in headers_txt.get(1.0, tk.END).strip().split('\n'):  # 取得標頭
        key, value = header.split(':')
        HEADERS[key] = value.strip()
    AVOID_URLS, SCAN_URLS = [], []  # 定義避免檢查的網址與想要檢查的網址
    for avoid_url in avoid_urls_txt.get(1.0, tk.END).strip().split('\n'):  # 取得避免檢查的網址
        AVOID_URLS.append(avoid_url.strip())
    for scan_url in scan_urls_txt.get(1.0, tk.END).strip().split('\n'):  # 取得想要檢查的網址
        SCAN_URLS.append(scan_url.strip())
    TIMEOUT = int(timeout_txt.get())  # 取得逾時秒數
    LAYER = int(layer_txt.get())  # 取得連結檢查層數
    ALT_MUST = True if alt_must_txt.get() == "True" else False  # 取得是否必須偵測圖片的 alt 屬性
    start_url_txt.delete(0, tk.END)  # 清空 start_url
    start_url_txt.insert(0, ",".join(SCAN_URLS) if SCAN_URLS else "")  # 顯示想要檢查的網址

    global yaml
    yaml.preserve_quotes = True
    yaml.indent(mapping=2, sequence=4, offset=2)
    with open('config.yaml', 'w', encoding='UTF-8') as f:
        yaml.dump(
            {
                'layer': LAYER,
                'timeout': TIMEOUT,
                'alt_must': ALT_MUST,
                'headers': HEADERS,
                'avoid_urls': AVOID_URLS,
                'scan_urls': SCAN_URLS,
            },
            f,
        )
    msg = "設定檔已儲存。"
    log_console.insert(tk.END, msg + "\n", "INFO")
    log_console.see(tk.END)  # 捲動 log_console 至最後一行


def clear_start_url():
    '''清除 start_url 的內容'''
    start_url_txt.delete(0, tk.END)  # 清空 start_url
    log_console.delete(1.0, tk.END)  # 清空 log_console


# 讀取設定檔 config.yaml
yaml = YAML()
yaml.preserve_quotes = True
yaml.indent(mapping=2, sequence=4, offset=2)
try:
    with open('config.yaml', 'r', encoding='UTF-8') as f:
        setting = yaml.load(f)
except FileNotFoundError:
    setting = dict()
    print("找不到設定檔 config.yaml，將使用預設設定...")
LAYER = setting.setdefault('layer', 3)  # 定義連結檢查層數
TIMEOUT = setting.setdefault('timeout', 8)  # 定義逾時秒數
ALT_MUST = setting.setdefault('alt_must', False)  # 定義是否必須偵測圖片的 alt 屬性
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6,zh-CN;q=0.5,la;q=0.4',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
    'Cookie': '_ga_SNR7NPLEYG=GS1.1.1651249603.1.0.1651250646.0; _ga_BGEHGPV3SB=GS1.1.1707539001.1.1.1707539323.0.0.0; _gid=GA1.3.1607128698.1707801742; _ga=GA1.1.381372473.1651249604; _ga_Q0EL30K2K5=GS1.1.1707801741.1.0.1707801770.0.0.0; _ga_54MVLT2EZN=GS1.1.1707843944.5.1.1707843969.0.0.0; __RequestVerificationToken_L05ldFNlcnZpY2Vz0=MrnqY4BqFXwyAR3uGWq5prQZPEwGWyzIJgIpuGFLyP8hqJ6eLKM9EWlC8NVA4MZqmyjtxmWT-9ZtzrO04NCTXcM_njimY7J0_WFWHlyWtzE1',
    'Sec-Ch-Ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
}
HEADERS = setting.setdefault('headers', headers)  # 定義連線標頭
avoid_urls = [
    "https://accessibility.moda.gov.tw/Applications/DetectLog/157716",
    "https://accessibility.moda.gov.tw/Home/Info/",
    "https://accessibility.moda.gov.tw/Download/Detail/1375?Category=52",
    "https://www.ndc.gov.tw/cp.aspx?n=32A75A78342B669D]",
]
AVOID_URLS = setting.setdefault('avoid_urls', avoid_urls)  # 定義避免檢查的網址
SCAN_URLS = setting.setdefault('scan_urls', ["https://www.ncut.edu.tw/"])  # 定義想要檢查的網址

# 系統設定
DOC_FOLDER = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Documents')  # 預設檔案存放目錄為【我的文件】
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # 關閉 SSL 不安全的警告訊息
logging.getLogger('requests').setLevel(logging.ERROR)  # 設定 requests 的 log 等級
logging.captureWarnings(True)  # 捕捉 py waring message
logger = None  # 定義全域變數
visited_link = dict()  # 存放已經檢查過的連結與狀態碼

# 建立主視窗
form = ttk.Window(themename="superhero")
form.title("網頁失效連結掃描工具")
form.geometry("1024x768")  # 設定視窗寬高
form.resizable(True, True)
# 指定行和列的權重
form.rowconfigure(0, weight=1)
form.columnconfigure(0, weight=1)
# 建立頁籤元件 Notebook
notebook = ttk.Notebook(form)
notebook.grid(row=0, column=0, padx=4, pady=4, sticky="nsew")

# 建立分頁
page1 = ttk.Frame(notebook)
page2 = ttk.Frame(notebook)
notebook.add(page1, text="網址掃描")
notebook.add(page2, text="系統設定")
page1.columnconfigure(0, weight=1)
page1.rowconfigure(1, weight=1)
page2.columnconfigure(0, weight=1)
page2.rowconfigure(1, weight=1)
page2.rowconfigure(2, weight=4)
page2.rowconfigure(3, weight=3)
# --- page1 ---
# 建立 frame1_1
frame1_1 = ttk.Frame(page1)
frame1_1.grid(row=0, column=0, padx=1, pady=1, sticky="nsew")
frame1_1.columnconfigure(1, weight=1)

# 建立網址輸入框Label
start_url_txt_label = ttk.Label(frame1_1, text="網址")
start_url_txt_label.grid(row=0, column=0, padx=5)

# 建立網址輸入框
start_url_txt = ttk.Entry(frame1_1)
start_url_txt.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
start_url_txt.insert(0, ",".join(SCAN_URLS) if SCAN_URLS else "")

# 建立右鍵選單
start_url_context_menu = tk.Menu(start_url_txt, tearoff=0)
start_url_context_menu.add_command(label="剪下", command=lambda: start_url_txt.event_generate("<<Cut>>"))
start_url_context_menu.add_command(label="複製", command=lambda: start_url_txt.event_generate("<<Copy>>"))
start_url_context_menu.add_command(label="貼上", command=lambda: start_url_txt.event_generate("<<Paste>>"))
start_url_context_menu.add_separator()
start_url_context_menu.add_command(
    label="刪除",
    command=lambda: start_url_txt.delete(tk.SEL_FIRST, tk.SEL_LAST) if start_url_txt.get() else None,
)
start_url_context_menu.add_command(label="清除", command=clear_start_url)
# 綁定右鍵事件
start_url_txt.bind("<Button-3>", lambda e: start_url_context_menu.post(e.x_root, e.y_root))

# 建立分析按鈕
analysis_btn = ttk.Button(
    frame1_1,
    text="掃描",
    command=lambda: analysis_func(start_url_txt.get(), LAYER),
    bootstyle="info",
    cursor='hand2',
)
Hovertip(analysis_btn, '按下按鈕開始掃描網址內的失效連結')
analysis_btn.grid(row=0, column=2, padx=5, sticky="e")

# 建立 frame1_2
frame1_2 = ttk.Frame(page1)
frame1_2.grid(row=1, column=0, padx=1, pady=1, sticky="nsew")
frame1_2.rowconfigure(0, weight=1)
frame1_2.columnconfigure(0, weight=1)

log_console = scrolledtext.ScrolledText(frame1_2, wrap=tk.WORD)
log_console.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
log_console.tag_config('ERROR', foreground='#e64530')  # 設定 ERROR 的文字顏色
log_console.tag_config('INFO', foreground='#c5c6c4')  # 設定 INFO 的文字顏色
log_console.tag_config('SUCCESS', foreground='#8ae234')  # 設定 SUCCESS 的文字顏色

# --- page2 ---
# 建立 frame2_1
frame2_1 = ttk.Frame(page2)
frame2_1.grid(row=0, column=0, padx=1, pady=1, sticky="nsew")
frame2_1.columnconfigure(6, weight=1)

# 建立連結的層數Label
layer_txt_label = ttk.Label(frame2_1, text="檢查連結的層數")
layer_txt_label.grid(row=0, column=0, padx=5)

# 建立連結的層數
layer_txt = ttk.Entry(frame2_1, width=3)
layer_txt.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
layer_txt.insert(0, LAYER)

# 建立連線逾時秒數Label
timeout_txt_label = ttk.Label(frame2_1, text="連線逾時秒數")
timeout_txt_label.grid(row=0, column=2, padx=5)

# 建立連線逾時秒數
timeout_txt = ttk.Entry(frame2_1, width=3)
timeout_txt.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
timeout_txt.insert(0, TIMEOUT)

# 建立圖片是否必須有 alt 屬性Label
alt_must_txt_label = ttk.Label(frame2_1, text="圖片是否必須有 alt 屬性")
alt_must_txt_label.grid(row=0, column=4, padx=5)

# 建立圖片是否必須有 alt 屬性
alt_must_txt = ttk.Combobox(frame2_1, values=["False", "True"], width=6)
alt_must_txt.grid(row=0, column=5, padx=5, pady=5, sticky="ew")
alt_must_txt.set(str(ALT_MUST))

cfg_save_btn = ttk.Button(frame2_1, text="儲存設定", command=save_config, bootstyle="success", cursor='hand2')
cfg_save_btn.grid(row=0, column=6, padx=5, sticky="e")

# 建立 frame2_2 用 LabelFrame
frame2_2 = ttk.Labelframe(page2, text="請求的標頭", style='info.TLabelframe')
frame2_2.grid(row=1, column=0, padx=3, pady=5, sticky="nsew")
frame2_2.columnconfigure(0, weight=1)
frame2_2.rowconfigure(0, weight=1)

# 建立想要避免檢查的網址清單
headers_txt = scrolledtext.ScrolledText(frame2_2, wrap=tk.WORD)
headers_txt.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
for header in HEADERS:
    headers_txt.insert(tk.END, f"{header}: {HEADERS[header]}\n")

# 建立 frame2_3 用 LabelFrame
frame2_3 = ttk.Labelframe(page2, text="想要避免檢查的網址清單", style='info.TLabelframe')
frame2_3.grid(row=2, column=0, padx=3, pady=5, sticky="nsew")
frame2_3.columnconfigure(0, weight=1)
frame2_3.rowconfigure(0, weight=1)

# 建立想要避免檢查的網址清單
avoid_urls_txt = scrolledtext.ScrolledText(frame2_3, wrap=tk.WORD)
avoid_urls_txt.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
for url in AVOID_URLS:
    avoid_urls_txt.insert(tk.END, f"{url}\n")

# 建立 frame2_4 用 LabelFrame
frame2_4 = ttk.Labelframe(page2, text="想要檢查的網址清單", style='success.TLabelframe')
frame2_4.grid(row=3, column=0, padx=3, pady=5, sticky="nsew")
frame2_4.columnconfigure(0, weight=1)
frame2_4.rowconfigure(0, weight=1)

# 建立想要避免檢查的網址清單
scan_urls_txt = scrolledtext.ScrolledText(frame2_4, wrap=tk.WORD)
scan_urls_txt.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
for url in SCAN_URLS:
    scan_urls_txt.insert(tk.END, f"{url}\n")

# 開始主迴圈
form.mainloop()
