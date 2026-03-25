import logging
import os
import re
import time
from collections import deque
from collections.abc import Callable
from dataclasses import dataclass, field
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import openpyxl
import requests
from bs4 import BeautifulSoup, UnicodeDammit
from openpyxl.styles import Alignment, Font, PatternFill
from ruamel.yaml import YAML
from selenium import webdriver
from selenium.common.exceptions import NoSuchDriverException, WebDriverException

EmitFunc = Callable[[str, str], None]
StopFunc = Callable[[], bool]
EMPTY_CONTENT_MARKER = "200 (內容為空)"
SKIPPED_MARKER = "200 (避免檢查的網址)"
HTTP_WARNING_MARKER = "且使用 http 協定並不安全"
SPA_SHELL_MARKER = "200 (疑似前端框架頁面)"
SOFT_404_MARKER = "200 (疑似 soft-404)"
SUSPICIOUS_REDIRECT_MARKER = "200 (疑似異常重定向)"


@dataclass
class ScanOptions:
    """掃描流程需要的設定參數。"""

    headers: dict
    avoid_urls: list[str]
    timeout: int
    alt_must: bool
    check_http: bool
    skip_visited: bool = True
    url_normalization: dict = field(default_factory=dict)
    download_link_rules: dict = field(default_factory=dict)
    soft_404_rules: dict = field(default_factory=dict)
    redirect_rules: dict = field(default_factory=dict)
    request_control: dict = field(default_factory=dict)
    normalized_avoid_urls: set[str] = field(init=False, repr=False)

    def __post_init__(self) -> None:
        self.normalized_avoid_urls = {
            _normalize_url_for_compare(url, self.url_normalization)
            for url in self.avoid_urls
            if str(url).strip()
        }


@dataclass
class ScanContext:
    """掃描期間共用的執行環境與狀態。"""

    logger: logging.Logger
    visited_link: dict
    browser: webdriver.Chrome
    emit: EmitFunc | None = None
    should_stop: StopFunc | None = None
    domain_last_request_at: dict = field(default_factory=dict)

    def push(self, level: str, message: str) -> None:
        """同時寫入 logger，必要時也回呼到 UI。"""
        if level == "error":
            self.logger.error(message)
        elif level == "warning":
            self.logger.warning(message)
        else:
            self.logger.info(message)
        if self.emit is not None:
            self.emit(level, message)

    def stopped(self) -> bool:
        """回傳目前是否已被要求停止掃描。"""
        return self.should_stop is not None and self.should_stop()


@dataclass(frozen=True)
class LinkCheckResult:
    """單一連結檢查結果。"""

    url: str
    link_text: str
    status: str
    issue_type: str
    content_kind: str


def create_webdriver() -> webdriver.Chrome:
    """建立 Selenium Chrome 瀏覽器實例。"""
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--log-level=3")
    try:
        browser = webdriver.Chrome(options=chrome_options)
        browser.implicitly_wait(20)
        return browser
    except (NoSuchDriverException, WebDriverException) as exc:
        raise RuntimeError(f"Selenium Manager 啟動失敗：{exc}") from exc


def create_logger(log_folder: str, filename: str, with_console: bool = False) -> logging.Logger:
    """建立 logger，必要時同步輸出到 console。"""
    logger = logging.getLogger(filename)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    formatter = logging.Formatter(
        "[%(asctime)s - %(levelname)s] %(message)s",
        datefmt="%Y%m%d %H:%M:%S",
    )

    if with_console:
        console_handle = logging.StreamHandler()
        console_handle.setLevel(logging.INFO)
        console_handle.setFormatter(formatter)
        logger.addHandler(console_handle)

    logname = os.path.join(log_folder, f"{filename}.log")
    file_handle = logging.FileHandler(logname, "w", "utf-8")
    file_handle.setLevel(logging.INFO)
    file_handle.setFormatter(formatter)
    logger.addHandler(file_handle)
    return logger


def close_logger(logger: logging.Logger) -> None:
    """關閉並移除 logger 的所有 handlers。"""
    for handler in logger.handlers[:]:
        handler.close()
        logger.removeHandler(handler)


def load_visited_link(visited_link_file: str) -> dict:
    """載入已檢查過且成功的連結清單。"""
    if os.path.exists(visited_link_file):
        yaml = YAML()
        yaml.preserve_quotes = True
        yaml.indent(mapping=2, sequence=4, offset=2)
        with open(visited_link_file, "r", encoding="utf-8") as file:
            data = yaml.load(file)
            return data or {}
    return {}


def save_visited_link(visited_link_file: str, visited_link: dict) -> None:
    """儲存已檢查過且成功的連結清單。"""
    yaml = YAML()
    yaml.preserve_quotes = True
    yaml.indent(mapping=2, sequence=4, offset=2)
    os.makedirs(os.path.dirname(visited_link_file) or ".", exist_ok=True)
    with open(visited_link_file, "w", encoding="utf-8") as file:
        for key, value in visited_link.items():
            if _is_cacheable_status(value):
                yaml.dump({key: value}, file)


def _requests_get(
    url: str,
    headers: dict,
    timeout: int,
    allow_redirects: bool = True,
    referer: str | None = None,
    ) -> requests.Response:
    """依協定包裝 requests.get。"""
    request_headers = dict(headers)
    if referer:
        request_headers["Referer"] = referer
    protocol = urlparse(url).scheme
    if protocol == "https":
        return requests.get(url, headers=request_headers, timeout=timeout, allow_redirects=allow_redirects)
    return requests.get(
        url,
        headers=request_headers,
        timeout=timeout,
        allow_redirects=allow_redirects,
        verify=False,
    )


def _wait_for_domain_slot(url: str, options: ScanOptions, context: ScanContext) -> None:
    """同一網域請求之間保留最小間隔，避免過度密集。"""
    delay = float(options.request_control.get("domain_delay_seconds", 0))
    if delay <= 0:
        return

    hostname = _normalized_hostname(url)
    if not hostname:
        return

    now = time.monotonic()
    last_request_at = context.domain_last_request_at.get(hostname)
    if last_request_at is not None:
        wait_seconds = delay - (now - last_request_at)
        if wait_seconds > 0:
            time.sleep(wait_seconds)
    context.domain_last_request_at[hostname] = time.monotonic()


def _request_with_retry(
    url: str,
    options: ScanOptions,
    context: ScanContext,
    allow_redirects: bool = True,
    referer: str | None = None,
) -> requests.Response:
    """加入 retry、backoff 與每網域節流的 requests 包裝。"""
    retry_count = int(options.request_control.get("retry_count", 0))
    backoff_seconds = float(options.request_control.get("backoff_seconds", 0))
    last_exception = None

    for attempt in range(retry_count + 1):
        _wait_for_domain_slot(url, options, context)
        try:
            response = _requests_get(
                url,
                options.headers,
                options.timeout,
                allow_redirects=allow_redirects,
                referer=referer,
            )
            if response.status_code not in {429, 500, 502, 503, 504} or attempt >= retry_count:
                return response
            last_exception = requests.exceptions.HTTPError(f"HTTP {response.status_code}")
        except (
            requests.exceptions.ConnectTimeout,
            requests.exceptions.ConnectionError,
            requests.exceptions.Timeout,
            requests.exceptions.SSLError,
        ) as exc:
            last_exception = exc
            if attempt >= retry_count:
                raise
        if attempt < retry_count:
            time.sleep(backoff_seconds * (2 ** attempt))

    if last_exception is not None:
        raise last_exception
    raise requests.exceptions.RequestException(f"無法取得此網頁內容：{url}")


def _is_problem_status(status: str) -> bool:
    """判斷狀態是否應被視為錯誤或警告。"""
    return (
        "200" not in status
        or EMPTY_CONTENT_MARKER in status
        or SOFT_404_MARKER in status
        or SUSPICIOUS_REDIRECT_MARKER in status
    )


def _is_cacheable_status(status: str) -> bool:
    """判斷狀態是否適合寫回成功連結快取。"""
    if "200" not in status:
        return False
    return not (
        EMPTY_CONTENT_MARKER in status
        or SKIPPED_MARKER in status
        or SOFT_404_MARKER in status
        or SUSPICIOUS_REDIRECT_MARKER in status
    )


def _classify_issue_type(status: str) -> str:
    """根據狀態字串回傳報表用錯誤類型。"""
    lowered = status.lower()
    if SKIPPED_MARKER in status:
        return "skipped"
    if EMPTY_CONTENT_MARKER in status:
        return "empty_content"
    if SOFT_404_MARKER in status:
        return "soft_404"
    if SUSPICIOUS_REDIRECT_MARKER in status:
        return "suspicious_redirect"
    if SPA_SHELL_MARKER in status:
        return "spa_shell"
    if HTTP_WARNING_MARKER in status:
        return "http_insecure"
    if "重新導向次數過多" in status:
        return "redirect_loop"
    if "連線逾時" in status:
        return "timeout"
    if "無法連線至此網頁" in status:
        return "connection_error"
    if "SSL 錯誤" in status:
        return "ssl_error"
    if "連結缺少協定" in status:
        return "invalid_url"
    if status.startswith("403"):
        return "http_403"
    if status.startswith("404"):
        return "http_404"
    if status.startswith("410"):
        return "http_410"
    if status.startswith("429"):
        return "http_429"
    if status.startswith("500"):
        return "http_500"
    if status.startswith("502"):
        return "http_502"
    if status.startswith("503"):
        return "http_503"
    if status.startswith("504"):
        return "http_504"
    if status.startswith("200 重定向"):
        return "redirect"
    if status.startswith("200"):
        return "ok"
    return "request_error"


def _is_html_like_url(url: str, options: ScanOptions | None = None) -> bool:
    """判斷網址是否看起來像可繼續爬的網頁。"""
    if options is not None and _is_probable_download_url(url, options):
        return False
    match = re.search(r".*\.\w{2,4}$", url)
    if not match:
        return True
    return url.endswith(("html", "htm", "php", "asp", "aspx", "jsp", "tw", "com"))


def _detect_content_kind(response: requests.Response, options: ScanOptions) -> str:
    """根據回應標頭判斷內容型態。"""
    content_disposition = (response.headers.get("Content-Disposition") or "").lower()
    if "attachment" in content_disposition:
        return "download"

    content_type = (response.headers.get("Content-Type") or "").split(";")[0].strip().lower()
    if content_type.startswith("text/html") or content_type in {"application/xhtml+xml"}:
        return "html"
    if content_type.startswith("image/"):
        return "image"
    if content_type.startswith("text/"):
        return "text"
    if content_type in {
        "application/pdf",
        "application/zip",
        "application/x-zip-compressed",
        "application/octet-stream",
        "application/msword",
        "application/vnd.ms-excel",
        "application/vnd.ms-powerpoint",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    }:
        return "download"
    if _is_probable_download_url(response.url, options):
        return "download"
    return "unknown"


def _is_local_http_target(url: str) -> bool:
    """判斷是否為本機測試用的 HTTP 網址。"""
    hostname = (urlparse(url).hostname or "").lower()
    return hostname in {"localhost", "127.0.0.1"}


def _normalized_hostname(url: str) -> str:
    """取得標準化後的主機名稱。"""
    return (urlparse(url).hostname or "").lower()


def _normalize_url_for_compare(url: str, rules: dict | None = None) -> str:
    """將網址轉成適合比對與快取的標準格式。"""
    parsed = urlparse(str(url).strip())
    if not parsed.scheme or not parsed.netloc:
        return str(url).strip()

    rules = rules or {}
    scheme = parsed.scheme.lower() if rules.get("lowercase_scheme_host") == "yes" else parsed.scheme
    hostname = parsed.hostname or ""
    if rules.get("lowercase_scheme_host") == "yes":
        hostname = hostname.lower()

    port = parsed.port
    if rules.get("strip_default_port") == "yes":
        if (scheme == "http" and port == 80) or (scheme == "https" and port == 443):
            port = None
    netloc = hostname
    if port:
        netloc = f"{hostname}:{port}"

    path = parsed.path or "/"
    if rules.get("collapse_slashes") == "yes":
        path = re.sub(r"/{2,}", "/", path)
    if rules.get("strip_trailing_slash") == "yes" and path not in {"", "/"}:
        path = path.rstrip("/") or "/"

    query_pairs = parse_qsl(parsed.query, keep_blank_values=True)
    remove_query_params = set(rules.get("remove_query_params") or [])
    if remove_query_params:
        query_pairs = [(key, value) for key, value in query_pairs if key.lower() not in remove_query_params]
    if rules.get("sort_query_params") == "yes":
        query_pairs = sorted(query_pairs)
    query = urlencode(query_pairs, doseq=True)
    fragment = "" if rules.get("drop_fragment") == "yes" else parsed.fragment

    return urlunparse((scheme, netloc, path, "", query, fragment))


def _get_visit_key(url: str, options: ScanOptions) -> str:
    """取得快取與去重用的網址 key。"""
    return _normalize_url_for_compare(url, options.url_normalization)


def _is_avoid_url(url: str, options: ScanOptions) -> bool:
    """判斷網址是否在略過清單中。"""
    normalized_url = _get_visit_key(url, options)
    return url in options.avoid_urls or normalized_url in options.normalized_avoid_urls


def _should_fallback_to_browser(status: str) -> bool:
    """判斷目前狀態是否需要改用 Selenium 補驗。"""
    return "連線逾時" in status or "無法連線至此網頁" in status


def _looks_like_spa_shell(response: requests.Response) -> bool:
    """判斷 HTML 是否看起來只是 Vue/React 類前端框架的殼。"""
    content_type = (response.headers.get("Content-Type") or "").lower()
    if "html" not in content_type:
        return False

    dammit = UnicodeDammit(response.content, ["utf-8", "latin-1", "iso-8859-1", "windows-1251"])
    markup = dammit.unicode_markup or ""
    if not markup.strip():
        return True

    soup = BeautifulSoup(markup, "lxml")
    body = soup.body or soup
    visible_text = body.get_text(" ", strip=True)
    visible_text = re.sub(r"\s+", " ", visible_text)
    script_count = len(soup.find_all("script"))
    has_mount_node = soup.select_one("#app, #root, #__next, #__nuxt, [data-reactroot], [data-react-checksum]") is not None
    has_framework_hint = any(
        token in markup.lower()
        for token in ("__vite__", "data-server-rendered", "id=\"app\"", "id=\"root\"", "id=\"__next\"", "id=\"__nuxt\"")
    )
    noscript_text = " ".join(tag.get_text(" ", strip=True).lower() for tag in soup.find_all("noscript"))
    asks_for_js = any(
        phrase in noscript_text
        for phrase in ("enable javascript", "javascript", "啟用 javascript", "需啟用 javascript", "請開啟 javascript")
    )

    return len(visible_text) <= 40 and (has_mount_node or has_framework_hint or asks_for_js or script_count >= 3)


def _extract_html_text(response: requests.Response) -> tuple[str, str, str]:
    """回傳 HTML 原始內容、標題與可見文字。"""
    dammit = UnicodeDammit(response.content, ["utf-8", "latin-1", "iso-8859-1", "windows-1251"])
    markup = dammit.unicode_markup or ""
    soup = BeautifulSoup(markup, "lxml")
    title = soup.title.get_text(" ", strip=True) if soup.title else ""
    body = soup.body or soup
    visible_text = re.sub(r"\s+", " ", body.get_text(" ", strip=True))
    return markup, title, visible_text


def _match_keywords(text: str, keywords: list[str]) -> list[str]:
    """找出文字中命中的關鍵字。"""
    lowered = text.lower()
    return [keyword for keyword in keywords if keyword and keyword in lowered]


def _detect_soft_404(response: requests.Response, options: ScanOptions) -> str | None:
    """判斷 200 頁面是否疑似 soft-404。"""
    rules = options.soft_404_rules
    if rules.get("enabled") != "yes":
        return None

    content_type = (response.headers.get("Content-Type") or "").lower()
    if "html" not in content_type:
        return None

    _markup, title, visible_text = _extract_html_text(response)
    final_url = response.url.lower()
    matched_title = _match_keywords(title, rules.get("title_keywords") or [])
    matched_body = _match_keywords(visible_text, rules.get("body_keywords") or [])
    matched_url = _match_keywords(final_url, rules.get("url_keywords") or [])
    hit_count = len(set(matched_title + matched_body + matched_url))

    if not matched_title and not matched_body:
        return None
    if len(visible_text) > int(rules.get("max_text_length", 1200)):
        return None
    if hit_count < int(rules.get("min_keyword_hits", 2)):
        return None

    reason_tokens = []
    if matched_title:
        reason_tokens.append(f"標題含 {matched_title[0]}")
    if matched_body:
        reason_tokens.append(f"內容含 {matched_body[0]}")
    if matched_url:
        reason_tokens.append(f"網址含 {matched_url[0]}")
    return "、".join(reason_tokens[:3])


def _classify_redirect(full_url: str, response: requests.Response, options: ScanOptions) -> str | None:
    """將重新導向結果整理成可讀的分類訊息。"""
    rules = options.redirect_rules
    if rules.get("classify_redirects") != "yes":
        return None

    if not response.history and _get_visit_key(full_url, options) == _get_visit_key(response.url, options):
        return None

    target = urlparse(response.url)
    history_code = response.history[-1].status_code if response.history else response.status_code
    categories = []
    suspicious_reason = None

    if _normalized_hostname(full_url) != _normalized_hostname(response.url):
        categories.append("站外")
        if rules.get("treat_cross_domain_as_warning") == "yes":
            suspicious_reason = "跨網域"
    else:
        categories.append("站內")

    target_path = (target.path or "/").lower()
    login_keywords = rules.get("suspicious_login_keywords") or []
    error_keywords = rules.get("suspicious_error_keywords") or []
    if any(keyword in target_path for keyword in login_keywords):
        suspicious_reason = "導向登入頁"
        categories.append("登入頁")
    elif any(keyword in target_path for keyword in error_keywords):
        suspicious_reason = "導向錯誤頁"
        categories.append("錯誤頁")

    category_text = "／".join(categories)
    if suspicious_reason:
        return f"{SUSPICIOUS_REDIRECT_MARKER}：{suspicious_reason}，{history_code} 到 {response.url}（{category_text}）"
    return f"200 重定向（{category_text}）：{history_code} 到 {response.url}"


def _is_probable_download_url(url: str, options: ScanOptions) -> bool:
    """判斷網址是否較像下載檔案而非可遞迴的 HTML 頁面。"""
    rules = options.download_link_rules
    parsed = urlparse(url)
    path = (parsed.path or "").lower()
    if "." in path.rsplit("/", 1)[-1]:
        extension = path.rsplit(".", 1)[-1]
        if extension in set(rules.get("extensions") or []):
            return True

    if any(keyword in path for keyword in (rules.get("path_keywords") or [])):
        return True

    query_pairs = parse_qsl(parsed.query, keep_blank_values=True)
    query_keys = set(rules.get("query_keys") or [])
    query_value_keywords = rules.get("query_value_keywords") or []
    for key, value in query_pairs:
        key_lower = key.lower()
        value_lower = value.lower()
        if key_lower in query_keys:
            return True
        if any(keyword in value_lower for keyword in query_value_keywords):
            return True
    return False


def _browser_fetch_status(full_url: str, context: ScanContext, original_status: str) -> str:
    """改用瀏覽器嘗試開啟網址，補強 requests 無法判斷的情況。"""
    try:
        context.browser.get(full_url)
        page_source = context.browser.page_source or ""
        body_text = page_source.strip()
        if not body_text:
            return f"200 改以瀏覽器自動化測試開啟成功，但內容為空，原訊息：{original_status}"
        return f"200 改以瀏覽器自動化測試開啟成功，原訊息：{original_status}"
    except WebDriverException:
        return original_status


def check_link(base_url: str, link: str, link_text: str, options: ScanOptions, context: ScanContext) -> LinkCheckResult:
    """檢查單一連結並回傳狀態字串。"""
    if link.startswith("http"):
        full_url = link.strip()
    else:
        full_url = urljoin(base_url, link).strip()
    visit_key = _get_visit_key(full_url, options)

    if _is_avoid_url(full_url, options):
        status = SKIPPED_MARKER
        content_kind = "skip"
    elif options.skip_visited and (visit_key in context.visited_link or full_url in context.visited_link):
        cached_status = context.visited_link.get(visit_key, context.visited_link.get(full_url))
        status = f"{cached_status} (已檢查過網址)"
        content_kind = "cached"
    else:
        protocol = urlparse(full_url).scheme
        try:
            response = _request_with_retry(full_url, options, context, allow_redirects=True, referer=base_url)
            status = str(response.status_code)
            content_length = len(response.content)
            content_kind = _detect_content_kind(response, options)
            if status == "200" and content_length == 0:
                status = EMPTY_CONTENT_MARKER
            elif status == "200":
                redirect_status = _classify_redirect(full_url, response, options)
                soft_404_reason = _detect_soft_404(response, options)
                if soft_404_reason:
                    status = f"{SOFT_404_MARKER}：{soft_404_reason}"
                elif content_kind == "html" and _looks_like_spa_shell(response):
                    status = SPA_SHELL_MARKER
                elif redirect_status:
                    status = redirect_status
        except requests.exceptions.ConnectTimeout as exc:
            status = f"連線逾時：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except requests.exceptions.ConnectionError as exc:
            status = f"無法連線至此網頁：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except requests.exceptions.Timeout as exc:
            status = f"連線逾時：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except requests.exceptions.MissingSchema as exc:
            status = f"連結缺少協定：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except requests.exceptions.TooManyRedirects as exc:
            status = f"重新導向次數過多：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except requests.exceptions.SSLError as exc:
            status = f"SSL 錯誤或憑證不正確：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except requests.exceptions.RequestException as exc:
            status = f"無法取得此網頁內容：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except requests.exceptions.HTTPError as exc:
            status = f"HTTP 錯誤：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"
        except Exception as exc:
            status = f"其它錯誤：{link}  錯誤訊息：{exc}"
            content_kind = "unknown"

        if _should_fallback_to_browser(status):
            status = _browser_fetch_status(full_url, context, status)
        elif status == SPA_SHELL_MARKER:
            status = _browser_fetch_status(full_url, context, status)

        if status == "403":
            status = f"{status} 請求被拒絕"
        elif status == "404":
            status = f"{status} 請求失敗"
        elif status == "410":
            status = f"{status} 請求的資源已經不存在"
        elif status == "500":
            status = f"{status} 伺服器錯誤"
        elif protocol == "http" and not _is_local_http_target(full_url):
            status = f"{status} {HTTP_WARNING_MARKER}"

        context.visited_link[visit_key] = status

    issue_type = _classify_issue_type(status)
    message = f"檢查: {link} ... 狀態: {status} ... 文字: {link_text}"
    if "200" in status:
        if (
            HTTP_WARNING_MARKER in status
            or EMPTY_CONTENT_MARKER in status
            or SOFT_404_MARKER in status
            or SUSPICIOUS_REDIRECT_MARKER in status
        ):
            context.push("warning", message)
        else:
            context.push("info", message)
    else:
        context.push("error", message)
    return LinkCheckResult(
        url=full_url,
        link_text=link_text,
        status=status,
        issue_type=issue_type,
        content_kind=content_kind,
    )


def get_links(url: str, options: ScanOptions, context: ScanContext) -> tuple[list, list, list, list]:
    """取得頁面中的內部連結、外部連結、缺少 alt 的圖片與 HTTP 連結。"""
    try:
        domain = _normalized_hostname(url)
        response = _request_with_retry(url.strip(), options, context, allow_redirects=True, referer=url)
        real_url = response.url
        real_domain = _normalized_hostname(real_url)
        if _get_visit_key(url, options) != _get_visit_key(real_url, options) and domain != real_domain:
            return ([], [], [], [])

        dammit = UnicodeDammit(response.content, ["utf-8", "latin-1", "iso-8859-1", "windows-1251"])
        soup = BeautifulSoup(dammit.unicode_markup, "lxml")
        result = soup.find("meta", attrs={"http-equiv": "refresh"})
        if result:
            _wait, client_url = result["content"].split(";")
            if client_url.strip().lower().startswith("url="):
                client_url = urljoin(url, client_url.strip()[4:])
            response = _request_with_retry(client_url, options, context, allow_redirects=True, referer=url)
            dammit = UnicodeDammit(response.content, ["utf-8", "latin-1", "iso-8859-1", "windows-1251"])
            soup = BeautifulSoup(dammit.unicode_markup, "lxml")
            context.push("info", f"網頁於前端重新導向至：{client_url}")

        all_links = []
        internal_links, external_links, no_alt_links, http_links = [], [], [], []

        for tag in soup.find_all(href=True):
            link = tag.get("href").strip()
            link_text = tag.text.strip() or link.split("/")[-1]
            all_links.append((link, link_text))

        for tag in soup.find_all(src=True):
            link = tag.get("src").strip()
            link_text = tag.get("alt", "").strip() or link.split("/")[-1]
            all_links.append((link, link_text))

        if options.alt_must:
            no_alt_links = [
                (tag.get("src"), tag.get("src").split("/")[-1]) for tag in soup.find_all("img", alt=False)
            ]

        for link, link_text in list(set(all_links)):
            absolute_link = urljoin(url, link).strip()
            link_domain = _normalized_hostname(absolute_link)
            scheme = urlparse(absolute_link).scheme
            if link.startswith(("#", "javascript", "mailto", "data:image")):
                continue
            if scheme not in ("http", "https"):
                continue
            if link_domain == domain:
                internal_links.append((absolute_link, link_text))
                if absolute_link.startswith("http://") and options.check_http and not _is_local_http_target(absolute_link):
                    http_links.append((absolute_link, link_text))
            elif link.startswith(("http", "//")):
                external_links.append((absolute_link, link_text))
                if absolute_link.startswith("http://") and options.check_http and not _is_local_http_target(absolute_link):
                    http_links.append((absolute_link, link_text))
            else:
                internal_links.append((absolute_link, link_text))
                if absolute_link.startswith("http://") and options.check_http and not _is_local_http_target(absolute_link):
                    http_links.append((absolute_link, link_text))

        return internal_links, external_links, no_alt_links, http_links
    except Exception as exc:
        message = f"無法取得此網頁內容：{url}  錯誤訊息：{exc}"
        context.push("error", message)
        return ([(url, f"無法取得此網頁內容：{exc}")], [], [], [])


def scan_site(start_url: str, depth_limit: int, options: ScanOptions, context: ScanContext) -> list[dict]:
    """使用佇列方式掃描指定網站並回傳結果。"""
    normalized_start_url = _get_visit_key(start_url, options)
    visited_url = set()
    queue = deque([(start_url, normalized_start_url, 0)])
    queued_url = {normalized_start_url}
    all_err_links = []
    start_domain = _normalized_hostname(start_url)

    while queue:
        if context.stopped():
            break

        url, normalized_url, current_depth = queue.popleft()
        queued_url.discard(normalized_url)
        if _is_avoid_url(url, options):
            continue
        if not _is_html_like_url(url, options):
            continue
        if normalized_url in visited_url or current_depth > depth_limit:
            continue

        context.push("info", f"第 {current_depth} 層連結： {url}")
        visited_url.add(normalized_url)

        internal_links, external_links, no_alt_links, http_links = get_links(url, options, context)
        if internal_links:
            context.push("info", "內部連結：" + ", ".join([link for link, _ in internal_links]))
        if external_links:
            context.push("info", "外部連結：" + ", ".join([link for link, _ in external_links]))
        if no_alt_links:
            context.push("info", "沒有 alt 屬性的連結：" + ", ".join([link for link, _ in no_alt_links]))
        if http_links:
            context.push("info", "HTTP 連結：" + ", ".join([link for link, _ in http_links]))
        internal_statuses = []
        for link, link_text in internal_links:
            if context.stopped():
                break
            result = check_link(url, link, link_text, options, context)
            internal_statuses.append(result)

        error_internal_links = [item for item in internal_statuses if _is_problem_status(item.status)]
        error_external_links = []
        error_links = error_internal_links + error_external_links

        if error_links:
            context.push("error", "錯誤連結：" + ", ".join([item.url for item in error_links]))

        if error_links or no_alt_links or http_links:
            all_err_links.append(
                {
                    "depth": current_depth,
                    "url": url,
                    "error_links": error_links,
                    "no_alt_links": no_alt_links,
                    "http_links": http_links,
                }
            )

        if current_depth >= depth_limit:
            continue

        error_internal_set = {item.url for item in error_internal_links}
        non_html_internal_set = {
            item.url for item in internal_statuses if item.content_kind in {"download", "image", "text"}
        }
        for link, link_text in internal_links:
            absolute_link = urljoin(url, link).strip()
            if absolute_link in error_internal_set:
                continue
            if absolute_link in non_html_internal_set:
                continue
            normalized_link = _get_visit_key(absolute_link, options)
            if _normalized_hostname(absolute_link) != start_domain:
                continue
            if _is_probable_download_url(absolute_link, options):
                continue
            if normalized_link not in visited_url and normalized_link not in queued_url:
                queue.append((absolute_link, normalized_link, current_depth + 1))
                queued_url.add(normalized_link)

    return all_err_links


def write_report(report_folder: str, filename: str, result: list[dict], include_http_links: bool) -> None:
    """將掃描結果寫入 Excel 報告。"""
    xlsx_name = os.path.join(report_folder, f"{filename}.xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.worksheets[0]
    sheet.append(("層數", "網頁", "錯誤類型", "錯誤連結", "連結文字", "狀態碼或錯誤訊息"))

    for column in range(1, sheet.max_column + 1):
        sheet.cell(1, column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(1, column).font = Font(bold=True)
    sheet.row_dimensions[1].height = 20

    for rec in sorted(result, key=lambda item: (item["depth"], item["url"])):
        for item in rec["error_links"]:
            sheet.append([rec["depth"], rec["url"], item.issue_type, item.url, item.link_text, item.status])
        for link, link_text in rec["no_alt_links"]:
            sheet.append([rec["depth"], rec["url"], "missing_alt", link, link_text, "圖片沒有 alt 屬性"])
        if include_http_links:
            for link, link_text in rec.get("http_links", []):
                sheet.append([rec["depth"], rec["url"], "http_insecure", link, link_text, "使用 http 協定並不安全"])

    column_widths = {"A": 8, "B": 50, "C": 18, "D": 70, "E": 35, "F": 70}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    for row in range(1, sheet.max_row + 1):
        sheet.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).style = "Hyperlink"
        sheet.cell(row, 2).hyperlink = sheet.cell(row, 2).value
        sheet.cell(row, 2).alignment = Alignment(vertical="center", wrap_text=True)

        sheet.cell(row, 4).style = "Hyperlink"
        if not sheet.cell(row, 4).value.startswith(("http", "//")):
            sheet.cell(row, 4).hyperlink = urljoin(sheet.cell(row, 2).value, sheet.cell(row, 4).value)
        else:
            sheet.cell(row, 4).hyperlink = sheet.cell(row, 4).value
        sheet.cell(row, 4).alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row, 5).alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row, 6).alignment = Alignment(vertical="center", wrap_text=True)

        for column in range(1, sheet.max_column + 1):
            sheet.cell(row, column).fill = (
                PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
                if row % 2 == 0
                else PatternFill(fill_type=None)
            )

    sheet.freeze_panes = sheet["A2"]
    workbook.save(xlsx_name)
