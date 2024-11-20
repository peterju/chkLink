import logging
import os
import re
import subprocess
import threading
import time
import tkinter as tk
from collections import deque
from datetime import datetime
from idlelib.tooltip import Hovertip
from tkinter import filedialog, messagebox, scrolledtext
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
import ttkbootstrap as ttk  # https://github.com/israel-dryer/ttkbootstrap
import urllib3
import wget
from bs4 import BeautifulSoup, UnicodeDammit
from openpyxl.styles import Alignment, Font, PatternFill
from py7zr import SevenZipFile
from ruamel.yaml import YAML
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.service import Service  # 配置和啟動 ChromeDriver 服務

# 更新 ChromeDriver ： https://googlechromelabs.github.io/chrome-for-testing/#stable
# from pprint import pprint
# from openpyxl.worksheet.hyperlink import Hyperlink
# https://titangene.github.io/article/python-logging.html


def stop_scanning():
    '''中斷掃描'''
    global stop_scan
    stop_scan = True


def run_update():
    '''更新程式'''

    def update_progress_bar(block_num, block_size, total_size):
        upd_progress_show.config(value=block_num, maximum=block_size)
        # if block_num == block_size:
        #     upd_progress_show.destroy()

    def wg():
        wget.download(
            'https://cc.ncut.edu.tw/var/file/32/1032/img/1517/RemoteVersion.yaml',
            bar=update_progress_bar,
        )
        with open('RemoteVersion.yaml', 'r', encoding='UTF-8') as f1:
            remote_ver = YAML().load(f1)
        with open('LocalVersion.yaml', 'r', encoding='UTF-8') as f2:
            local_ver = YAML().load(f2)

        if local_ver.get('version') >= remote_ver.get('version'):
            os.remove('RemoteVersion.yaml')
            messagebox.showwarning("資訊", "您的程式版本已是最新，故無需更新！")
        else:
            filename = wget.download(
                'https://cc.ncut.edu.tw/var/file/32/1032/img/1517/update.7z',
                bar=update_progress_bar,
            )
            with SevenZipFile(filename, 'r') as archive:
                archive.extractall()  # 解壓縮
            os.remove(filename)  # 刪除下載的壓縮檔案
            os.replace('RemoteVersion.yaml', 'LocalVersion.yaml')  # 更新版本記錄
            subprocess.run(["update.cmd"])  # 啟動更新程式

    if os.path.isfile('chklink_upd.exe'):
        os.remove('chklink_upd.exe')
    dl_thread = threading.Thread(target=wg)
    dl_thread.start()


def dl_resources():
    '''更新必要的元件'''
    URL = 'https://cc.ncut.edu.tw/var/file/32/1032/img/1517/resources.7z'
    filename = wget.download(URL)  # 呼叫 wget 下載檔案
    with SevenZipFile(filename, 'r') as archive:
        archive.extractall()  # 解壓縮
    os.remove(filename)  # 刪除檔案


def is_valid_link(base_url, link, link_text="") -> bool:
    '''檢查連結是否有效'''
    global logger, visited_link, browser

    if link.startswith("http"):
        full_url = link.strip()  # 如果連結已經是絕對連結，則直接使用
    else:
        full_url = urljoin(base_url, link).strip()  # 如果連結是相對連結，則加上基本連結後使用

    if full_url in AVOID_URLS:  # 若連結為避免檢查的網址
        status = "200 (避免檢查的網址)"
    elif full_url in visited_link and check_var.get() == 1:  # 若畫面上跳過檢查過的網址打勾，則跳過檢查
        status = f"{visited_link[full_url]} (已檢查過網址)"
    else:
        protocol = urlparse(full_url).scheme  # 取得協定
        try:
            if protocol == "https":
                response = requests.get(
                    full_url,
                    headers=HEADERS,
                    timeout=int(timeout_txt.get()),
                    allow_redirects=True,
                )  # 發送 GET 請求，其實發 HEAD 請求會比較好，但會返回 403 很奇怪，可能是 Server 端設定錯誤
            else:
                response = requests.get(
                    full_url,
                    headers=HEADERS,
                    timeout=int(timeout_txt.get()),
                    allow_redirects=True,
                    verify=False,
                )
            status = str(response.status_code)  # 取得狀態碼
            real_url = response.url  # 取得網頁的實際連結, 避免重定向造成誤判
            domain1 = urlparse(full_url).netloc  # 取得原始連結的網域
            domain2 = urlparse(real_url).netloc  # 取得實際連結的網域B
            if status == '200' and domain1 != domain2:  # 若原始連結與實際連結的網域不同(例如: 302 暫時定向到別的網域)
                if protocol == "https":
                    response = requests.get(
                        full_url, headers=HEADERS, timeout=int(timeout_txt.get()), allow_redirects=False
                    )
                else:
                    response = requests.get(
                        full_url,
                        headers=HEADERS,
                        timeout=int(timeout_txt.get()),
                        allow_redirects=False,
                        verify=False,
                    )
                status = f'200 重定向：{str(response.status_code)} 到 {real_url}'  # 設定狀態
        except requests.exceptions.ConnectTimeout as e:
            status = f"連線逾時：{link}  錯誤訊息：{e}"
        except requests.exceptions.ConnectionError as e:
            status = f"無法連線至此網頁：{link}  錯誤訊息：{e}"
        except requests.exceptions.Timeout as e:
            status = f"連線逾時：{link}  錯誤訊息：{e}"
        except requests.exceptions.MissingSchema as e:
            status = f"連結缺少協定：{link}  錯誤訊息：{e}"
        except requests.exceptions.TooManyRedirects as e:
            status = f"重新導向次數過多：{link}  錯誤訊息：{e}"
        except requests.exceptions.SSLError as e:
            status = f"SSL 錯誤或憑證不正確：{link}  錯誤訊息：{e}"
        except requests.exceptions.RequestException as e:
            status = f"無法取得此網頁內容：{link}  錯誤訊息：{e}"
        except requests.exceptions.HTTPError as e:
            status = f"HTTP 錯誤：{link}  錯誤訊息：{e}"
        except Exception as e:
            status = f"其它錯誤：{link}  錯誤訊息：{e}"
        if '連線逾時' in status or '無法連線至此網頁' in status:
            try:
                browser.get(full_url)  # 用瀏覽器打開網頁
                status = f"200 改以瀏覽器自動化測試開啟成功，原訊息：{status}"
            except WebDriverException:
                pass
        if status == "403":
            status = f"{status} 請求被拒絕"
        elif status == "404":
            status = f"{status} 請求失敗"
        elif status == "410":
            status = f"{status} 請求的資源已經不存在"
        elif status == "500":
            status = f"{status} 伺服器錯誤"
        elif protocol == "http":
            status = f"{status} 且使用 http 協定並不安全"
        visited_link[full_url] = status  # 將連結與狀態碼加入已檢查的集合

    msg = f"檢查: {link} ... 狀態: {status} ... 文字: {link_text}"
    if "200" in status:
        if "使用 http 協定並不安全" in status:
            logger.warning(msg)
            log_console.insert(tk.END, msg + "\n", "WARNING")
        else:
            logger.info(msg)
            log_console.insert(tk.END, msg + "\n", "INFO")
    else:
        logger.error(msg)
        log_console.insert(tk.END, msg + "\n", "ERROR")
    log_console.see(tk.END)  # 捲動 log_console 至最後一行
    return status


def get_links(url) -> tuple:
    '''取得網頁中的所有內部連結、外部連結、沒有 alt 的連結與 HTTP 連結'''
    global logger
    try:
        protocol = urlparse(url).scheme  # 取得協定
        domain = urlparse(url).netloc  # 取得網域
        if protocol == "https":
            response = requests.get(
                url.strip(), headers=HEADERS, timeout=int(timeout_txt.get()), allow_redirects=True
            )  # 發送 GET 請求
        else:
            response = requests.get(
                url.strip(), headers=HEADERS, timeout=int(timeout_txt.get()), allow_redirects=True, verify=False
            )
        real_url = response.url  # 取得網頁的實際連結, 避免重定向造成誤判
        real_domain = urlparse(real_url).netloc  # 取得網域
        if domain != real_domain:  # 若原始連結與實際連結的網域不同
            return ([], [], [], [])  # 回傳空串列
        # 使用 UnicodeDammit 來處理編碼問題
        dammit = UnicodeDammit(response.content, ["utf-8", "latin-1", "iso-8859-1", "windows-1251"])
        soup = BeautifulSoup(dammit.unicode_markup, 'lxml')  # 使用 BeautifulSoup 解析網頁內容
        result = soup.find("meta", attrs={"http-equiv": "refresh"})  # 取得 refresh 標籤
        if result:  # 若有 refresh 標籤，代表網頁有前端重新導向
            wait, client_url = result["content"].split(";")  # 取得重新導向的秒數與網址
            if client_url.strip().lower().startswith("url="):  # 若網址以 url= 開頭
                client_url = urljoin(url, client_url.strip()[4:])  # 取得絕對網址
            if protocol == "https":
                response = requests.get(
                    client_url, headers=HEADERS, timeout=int(timeout_txt.get()), allow_redirects=True
                )  # 發送 GET 請求
            else:
                response = requests.get(
                    client_url, headers=HEADERS, timeout=int(timeout_txt.get()), allow_redirects=True, verify=False
                )
            dammit = UnicodeDammit(response.content, ["utf-8", "latin-1", "iso-8859-1", "windows-1251"])
            soup = BeautifulSoup(dammit.unicode_markup, 'lxml')  # 使用 BeautifulSoup 解析網頁內容
            msg = f"網頁於前端重新導向至：{client_url}"
            logger.info(msg)
            log_console.insert(tk.END, msg + "\n", "INFO")
            log_console.see(tk.END)
        all_links = []  # 取得所有的連結
        internal_links, external_links, no_alt_links, http_links = (
            [],
            [],
            [],
            [],
        )  # 初始化內部連結、外部連結、沒有 alt 的連結與 HTTP 連結

        for tag in soup.find_all(href=True):  # 找到所有具有 href 屬性的標籤
            link = tag.get('href').strip()
            link_text = tag.text.strip() or link.split('/')[-1]  # 使用連結文字或檔名
            all_links.append((link, link_text))

        for tag in soup.find_all(src=True):  # 找到所有具有 src 屬性的標籤
            link = tag.get('src').strip()
            link_text = tag.get('alt', '').strip() or link.split('/')[-1]  # 使用 alt 屬性或檔名
            all_links.append((link, link_text))

        if alt_must_txt.get().lower() == 'yes':  # 若須偵測圖片的 alt 屬性
            no_alt_links = [
                (tag.get('src'), tag.get('src').split('/')[-1]) for tag in soup.find_all('img', alt=False)
            ]  # 找到所有圖片沒有 alt 屬性的標籤

        all_links = list(set(all_links))  # 去除重複的連結
        for link, link_text in all_links:
            link_domain = urlparse(link).netloc  # 取得連結的網域
            if link.startswith(
                ('#', 'javascript', 'mailto', 'data:image')
            ):  # 若連結為錨點、javascript、郵件連結或 base64圖片，則跳過
                continue
            elif link_domain == domain:  # 若連結為目前網址
                internal_links.append((link, link_text))
            elif link.startswith(('http', '//')):  # 若連結為 http 或 // 開頭的為外部連結
                external_links.append((link, link_text))
                if link.startswith('http://') and check_http_txt.get().lower() == 'yes':  # 若連結為 http 協定且需要檢查
                    http_links.append((link, link_text))
            elif link.startswith('/'):  # 若連結為 / 開頭的為內部連結
                internal_links.append((link, link_text))
            else:  # 其它的為內部連結，例如檔名開頭的
                internal_links.append((link, link_text))
        return (
            internal_links,
            external_links,
            no_alt_links,
            http_links,
        )  # 回傳內部連結、外部連結、圖片沒有 alt 屬性的連結與 HTTP 連結
    except Exception as e:
        msg = f"無法取得此網頁內容：{url}  錯誤訊息：{e}"
        logger.error(msg)
        log_console.insert(tk.END, msg + "\n", "ERROR")
        log_console.see(tk.END)  # 捲動 log_console 至最後一行
        # 將錯誤資訊加入到 internal_links 中
        internal_links = [(url, f"無法取得此網頁內容：{e}")]
        return (internal_links, [], [], [])  # 若發生錯誤，回傳包含錯誤資訊的串列


def analysis_func(start_urls, depth_limit=1):
    global stop_scan
    stop_scan = False  # 重置中斷變數
    log_console.delete(1.0, tk.END)  # 清空 log_console
    if len(start_urls) > 0:
        analysis_btn.config(state=tk.DISABLED, cursor='X_cursor')  # 按鈕設定為無效
        stop_btn.config(state=tk.NORMAL, cursor='hand2')  # 啟用中斷按鈕
        threading.Thread(
            target=queued_link_check, args=(start_urls, depth_limit)
        ).start()  # 使用 threading 在背景執行 queued_link_check 函式
    else:
        messagebox.showwarning("資訊", "請輸入網址才能進行失效連結掃描！")


def queued_link_check(start_urls, depth_limit=1) -> list:
    '''使用雙向佇列結構儲存網站中的連結'''
    global logger, HEADERS, AVOID_URLS, browser, stop_scan

    chrome_options = webdriver.ChromeOptions()  # 建立 Chrome 選項
    chrome_options.add_argument("--headless=new")  # 隱藏瀏覽器
    chrome_options.add_argument(
        "--log-level=3"
    )  # 隱藏錯誤訊息 INFO = 0, WARNING = 1, LOG_ERROR = 2, LOG_FATAL = 3; default is 0
    chrome_service = Service(executable_path='webdriver/chromedriver.exe')  # ChromeDriver 路徑
    browser = webdriver.Chrome(service=chrome_service, options=chrome_options)  # 建立 Chrome 瀏覽器實例
    browser.implicitly_wait(20)  # 隱性等待 20 秒

    HEADERS = dict()
    for header in headers_txt.get(1.0, tk.END).strip().split('\n'):  # 取得標頭
        key, value = header.split(':')
        HEADERS[key] = value.strip()
    AVOID_URLS = []  # 定義避免檢查的網址
    for avoid_url in avoid_urls_txt.get(1.0, tk.END).strip().split('\n'):  # 取得避免檢查的網址
        AVOID_URLS.append(avoid_url.strip())

    start_time = datetime.now()  # 取得開始時間
    for start_url in start_urls.split(','):  # 以逗號分隔多個網址
        if stop_scan:
            break  # 中斷掃描
        start_url = start_url.strip()  # 去除前後空白
        url_domain = urlparse(start_url).netloc  # 取得網域
        current_time = datetime.now()  # 取得目前時間
        filename = f"{url_domain}_{current_time.strftime('%m%d_%H%M')}"  # 建立檔案名稱，格式為 domain_月日_時分
        logger = create_logger(report_dir_txt.get(), filename)  # 建立 logger
        msg = f"=》開始掃描 {url_domain}..."
        logger.info(msg)
        log_console.insert(tk.END, msg + "\n", "INFO")
        log_console.see(tk.END)  # 捲動 log_console 至最後一行

        visited_url = set()  # 存放已經檢查過的連結
        queue = deque([(start_url, 0)])  # 存放待檢查的連結，設定目前深度為 0
        all_err_links = []  # 存放所有錯誤連結
        while queue:
            if stop_scan:
                break  # 中斷掃描
            url, current_depth = queue.popleft()  # 取得待檢查的連結
            if url in AVOID_URLS:  # 若連結為避免檢查的網址
                continue  # 跳過
            match = re.search(r".*\.\w{2,4}$", url)  # 是否符合具有副檔名的.xxx或.xxxx (例如: .7z, .php, .aspx)
            if match and not url.endswith(
                ('html', 'htm', 'php', 'asp', 'aspx', 'jsp', 'tw', 'com')
            ):  # 若連結為檔案且非網頁檔
                continue  # 跳過

            if url not in visited_url and current_depth <= depth_limit:  # 若連結未檢查過且深度未達到指定深度
                msg = f"第 {current_depth} 層連結： {url}"
                logger.info(msg)  # 顯示目前處理的連結
                log_console.insert(tk.END, msg + "\n", "INFO")
                log_console.see(tk.END)  # 捲動 log_console 至最後一行
                visited_url.add(url)  # 將連結加入已檢查的集合
                internal_links, external_links, no_alt_links, http_links = get_links(
                    url
                )  # 取得內部連結、外部連結、沒有alt的連結與 HTTP 連結
                if internal_links:
                    msg = f"內部連結：{', '.join([link for link, _ in internal_links])}"
                    logger.info(msg)
                    log_console.insert(tk.END, msg + "\n", "INFO")
                if external_links:
                    msg = f"外部連結：{', '.join([link for link, _ in external_links])}"
                    logger.info(msg)
                    log_console.insert(tk.END, msg + "\n", "INFO")
                if no_alt_links:
                    msg = f"沒有 alt 屬性的連結：{', '.join([link for link, _ in no_alt_links])}"
                    logger.info(msg)
                    log_console.insert(tk.END, msg + "\n", "INFO")
                if http_links:
                    msg = f"HTTP 連結：{', '.join([link for link, _ in http_links])}"
                    logger.info(msg)
                    log_console.insert(tk.END, msg + "\n", "INFO")
                log_console.see(tk.END)  # 捲動 log_console 至最後一行
                error_internal_links = []  # 存放錯誤的內部連結
                for link, link_text in internal_links:
                    if stop_scan:
                        break  # 中斷掃描
                    status = is_valid_link(url, link, link_text)  # 檢查連結是否有效
                    if '200' not in status:  # 若狀態碼不是 200
                        error_internal_links.append((link, status, link_text))  # 將錯誤的連結加入存放錯誤的內部連結
                error_external_links = []  # 存放錯誤的外部連結
                for link, link_text in external_links:  # 檢查外部連結是否有效
                    if stop_scan:
                        break  # 中斷掃描
                    status = is_valid_link(url, link, link_text)  # 檢查連結是否有效
                    if '200' not in status:  # 若狀態碼不是 200
                        error_external_links.append((link, status, link_text))  # 將錯誤的連結加入存放錯誤的外部連結
                error_links = []  # 存放此次所有 url 錯誤的連結
                # 將錯誤的內部連結加入所有錯誤連結集合
                error_links.extend(error_internal_links) if error_internal_links else error_links
                # 將錯誤的外部連結加入所有錯誤連結集合
                error_links.extend(error_external_links) if error_external_links else error_links
                error_links = list(set(error_links))  # 去除重複的錯誤連結
                if error_links:  # 若有錯誤連結存在
                    msg = f"錯誤連結：{', '.join([link for link, status, _ in error_links])}"
                    logger.error(msg)
                    log_console.insert(tk.END, msg + "\n", "ERROR")
                    log_console.see(tk.END)  # 捲動 log_console 至最後一行
                    for link, _ in internal_links:  # 將錯誤連結從內部連結中移除
                        for err_link, status, _ in error_links:
                            if (link, link_text) in internal_links:
                                internal_links.remove((link, link_text))

                if error_links or no_alt_links or http_links:
                    record = {  # 建立錯誤連結記錄
                        'depth': current_depth,
                        'url': url,
                        'error_links': error_links,
                        'no_alt_links': no_alt_links,
                        'http_links': http_links,  # 新增 HTTP 連結記錄
                    }
                    all_err_links.append(record)  # 將錯誤連結記錄加入全體錯誤連結記錄中

                # 將正確的內部連結加入待檢查的連結
                for link, link_text in internal_links:
                    if (
                        link and link.startswith((start_url, '/')) and current_depth <= depth_limit
                    ):  # 若連結為內部連結且深度未達到指定深度
                        absolute_link = urljoin(url, link).strip()  # 取得絕對連結, 並去除前後空白
                        queue.append((absolute_link, current_depth + 1))  # 將絕對連結加入待檢查的連結
        if all_err_links:
            report(report_dir_txt.get(), filename, all_err_links)  # 產生報告
            err_count = sum([len(rec['error_links']) for rec in all_err_links])
            msg = f"=》掃描 {url_domain} 完成，共有 {err_count} 個錯誤連結。"
            logger.info(msg)
            log_console.insert(tk.END, msg + "\n", "SUCCESS")
            log_console.see(tk.END)  # 捲動 log_console 至最後一行
            msg = f"=》報告已存放於 {report_dir_txt.get()}..."
            logger.info(msg)
            log_console.insert(tk.END, msg + "\n", "SUCCESS")
            log_console.see(tk.END)  # 捲動 log_console 至最後一行
        else:
            msg = f"=》太棒了！{url_domain} 沒有錯誤連結。"
            logger.info(msg)
            log_console.insert(tk.END, msg + "\n", "SUCCESS")
            log_console.see(tk.END)  # 捲動 log_console 至最後一行

    # 掃描結束
    browser.quit()  # 關閉瀏覽器
    save_visted_link()  # 儲存已檢查過的連結

    analysis_btn.config(state=tk.NORMAL, cursor='hand2')  # 完成後重新啟用分析按鈕
    stop_btn.config(state=tk.DISABLED, cursor='X_cursor')  # 禁用中斷按鈕
    end_time = datetime.now()  # 取得結束時間
    hours, remainder = divmod((end_time - start_time).seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    msg = f"=》全部掃描完成！共花費 {hours} 小時 {minutes} 分 {seconds} 秒。"
    logger.info(msg)
    log_console.insert(tk.END, msg + "\n", "SUCCESS")
    log_console.see(tk.END)  # 捲動 log_console 至最後一行

    # 移除 logger 所有的 handler
    for handler in logger.handlers[:]:
        handler.close()
        logger.removeHandler(handler)

    answer = messagebox.askquestion("資訊", "掃描完成！是否要開啟檔案所在目錄？")
    if answer == 'yes':
        subprocess.Popen(f'explorer "{report_dir_txt.get()}"')


def report(report_folder, filename, result) -> None:
    '''產生xlsx報告'''
    xlsx_name = os.path.join(report_folder, f"{filename}.xlsx")  # 報告檔路徑
    workbook = openpyxl.Workbook()  # 利用 Workbook 建立一個新的工作簿
    sheet = workbook.worksheets[0]  # 取得第一個工作表
    field_names = ('層數', '網頁', '錯誤連結', '連結文字', '狀態碼或錯誤訊息')  # 欄位名稱
    sheet.append(field_names)  # 寫入欄位名稱
    for column in range(1, sheet.max_column + 1):  # 設定第一列的第1欄到最後1欄
        sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, column).font = Font(bold=True)  # 設定粗體
    sheet.row_dimensions[1].height = 20  # 設定第一列高度
    for rec in result:
        for link, status, link_text in rec['error_links']:
            sheet.append([rec['depth'], rec['url'], link, link_text, status])  # 寫入一列 5 個欄位
        for link in rec['no_alt_links']:
            sheet.append([rec['depth'], rec['url'], link, link.split('/')[-1], "圖片沒有 alt 屬性"])
        if check_http_txt.get().lower() == 'yes':
            for link, link_text in rec.get('http_links', []):
                sheet.append([rec['depth'], rec['url'], link, link_text, "使用 http 協定並不安全"])
    # 設定欄寬
    column_widths = {'A': 8, 'B': 60, 'C': 70, 'D': 35, 'E': 20}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    for row in range(1, sheet.max_row + 1):  # 設定第一列到最後一列
        sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')  # 第1欄設定水平置中、垂直置中
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).style = "Hyperlink"  # 設定超連結樣式
        sheet.cell(row, 2).hyperlink = sheet.cell(row, 2).value
        sheet.cell(row, 2).alignment = Alignment(vertical='center', wrap_text=True)  # 第2欄設定垂直置中自動換行

        sheet.cell(row, 3).style = "Hyperlink"
        if not sheet.cell(row, 3).value.startswith(('http', '//')):
            sheet.cell(row, 3).hyperlink = urljoin(sheet.cell(row, 2).value, sheet.cell(row, 3).value)
        else:
            sheet.cell(row, 3).hyperlink = sheet.cell(row, 3).value
        sheet.cell(row, 3).alignment = Alignment(vertical='center', wrap_text=True)  # 第3欄設定垂直置中自動換行

        sheet.cell(row, 4).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True
        )  # 第4欄設定水平靠左、垂直置中、自動換行
        sheet.cell(row, 5).alignment = Alignment(vertical='center')  # 第5欄設定垂直置中

        for column in range(1, sheet.max_column + 1):
            sheet.cell(row, column).fill = (
                PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
                if row % 2 == 0
                else PatternFill(fill_type=None)
            )  # 設定偶數列的背景顏色
    sheet.freeze_panes = sheet['A2']  # 設定凍結第一列
    workbook.save(xlsx_name)  # 儲存檔案


def create_logger(log_folder, filename) -> logging.Logger:
    """建立 logger"""
    global logger, log_console
    logger = logging.getLogger(filename)  # 取得 logger
    logger.setLevel(logging.INFO)  # 設定日誌記錄等級(共有：DEBUG,INFO,WARNING,ERROR,CRITICAL)
    formatter = logging.Formatter(
        '[%(asctime)s - %(levelname)s] %(message)s',
        datefmt='%Y%m%d %H:%M:%S',
    )  # 設定 log 格式為：[日期時間 - 等級] 日誌訊息，其中的日期時間格式為 年月日 時分秒、等級為 INFO...5種

    # 設定檔案寫入
    logname = os.path.join(log_folder, f"{filename}.log")  # log 檔路徑
    file_handle = logging.FileHandler(logname, 'w', 'utf-8')
    file_handle.setLevel(logging.INFO)
    file_handle.setFormatter(formatter)

    # 為 log 綁定檔案
    logger.addHandler(file_handle)
    return logger


def choose_dir():
    '''更改分析後產生的文件目錄'''
    folder_selected = filedialog.askdirectory(initialdir='shell:personal', title="請選擇下載目錄")
    if folder_selected:  # 若有選擇目錄
        dl_folder = folder_selected.replace("/", "\\")  # 將路徑中的 / 替換為 \\
        report_dir_txt.delete(0, tk.END)  # 清空下載目錄 report_dir_txt
        report_dir_txt.insert(0, dl_folder)  # 將選擇的目錄顯示在 report_dir_txt


def clear_start_url():
    '''清除 start_url 的內容'''
    start_url_txt.delete(1.0, tk.END)  # 清空 start_url
    log_console.delete(1.0, tk.END)  # 清空 log_console


def save_config() -> None:
    '''儲存設定檔'''
    global HEADERS, AVOID_URLS, SCAN_URLS

    HEADERS = dict()  # 連線標頭
    for header in headers_txt.get(1.0, tk.END).strip().split('\n'):  # 取得標頭欄位值
        key, value = header.split(':')
        HEADERS[key] = value.strip()
    AVOID_URLS = []  # 避免檢查的網址
    for avoid_url in avoid_urls_txt.get(1.0, tk.END).strip().split('\n'):  # 取得避免檢查的網址欄位值
        AVOID_URLS.append(avoid_url.strip())
    SCAN_URLS = []  # 想要檢查的網址
    for scan_url in scan_urls_txt.get(1.0, tk.END).strip().split('\n'):  # 取得想要檢查的網址欄位值
        SCAN_URLS.append(scan_url.strip())
    start_url_txt.delete(1.0, tk.END)  # 清空 start_url
    start_url_txt.insert(1.0, ",".join(SCAN_URLS) if SCAN_URLS else "")  # 顯示想要檢查的網址

    yaml = YAML()
    yaml.preserve_quotes = True
    yaml.indent(mapping=2, sequence=4, offset=2)

    # 讀取現有的設定檔，保留註解
    with open(config_file, 'r', encoding='UTF-8') as f:
        current_config = yaml.load(f)

    # 更新設定值
    current_config['layer'] = int(layer_txt.get())  # 取得連結檢查層數
    current_config['timeout'] = int(timeout_txt.get())  # 取得逾時秒數
    current_config['alt_must'] = alt_must_txt.get()  # 取得是否必須偵測圖片的 alt 屬性
    current_config['check_http'] = check_http_txt.get()  # 取得是否檢查 HTTP 協定
    current_config['rpt_folder'] = report_dir_txt.get()  # 報告檔案目錄
    current_config['headers'] = HEADERS  # 連線標頭
    current_config['avoid_urls'] = AVOID_URLS  # 避免檢查的網址
    current_config['scan_urls'] = SCAN_URLS  # 想要檢查的網址

    # 將更新後的設定寫回檔案，保留註解
    with open(config_file, 'w', encoding='UTF-8') as f:
        yaml.dump(current_config, f)

    msg = "設定檔已儲存。"
    log_console.insert(tk.END, msg + "\n", "INFO")
    log_console.see(tk.END)  # 捲動 log_console 至最後一行


def save_visted_link() -> None:
    '''儲存檢查過但回應OK的連結'''
    global visited_link, visted_link_file
    yaml = YAML()  # 建立 YAML 物件
    yaml.preserve_quotes = True  # 保留引號
    yaml.indent(mapping=2, sequence=4, offset=2)  # 設定縮排
    with open(visted_link_file, 'w', encoding='UTF-8') as f:
        for k, v in visited_link.items():  # 從字典逐一取得項目(k=已檢查過的連結, v=回應狀態碼)
            if '200' in v:  # 若回應狀態碼為 200
                yaml.dump({k: v}, f)  # 寫入檔案


def load_visted_link(visted_link_file) -> dict:
    '''載入已檢查過的連結'''
    if os.path.exists(visted_link_file):
        yaml = YAML()  # 建立 YAML 物件
        yaml.preserve_quotes = True  # 保留引號
        yaml.indent(mapping=2, sequence=4, offset=2)  # 設定縮排
        with open(visted_link_file, 'r', encoding='UTF-8') as f:
            visited_link = yaml.load(f)  # 載入已檢查過的連結成為字典
    else:
        visited_link = dict()  # 若檔案不存在，則建立空字典
    return visited_link


def create_config(cfg_file) -> dict:
    # 建立設定檔 config.yaml
    yaml = YAML()
    yaml.preserve_quotes = True
    yaml.indent(mapping=2, sequence=4, offset=2)
    with open(cfg_file, 'w', encoding='UTF-8') as f:
        setting = dict()
        setting['layer'] = 3  # 定義連結檢查層數
        setting['timeout'] = 8  # 定義逾時秒數
        setting['alt_must'] = 'no'  # 定義是否必須偵測圖片的 alt 屬性
        setting['check_http'] = 'yes'  # 定義是否檢查 HTTP 協定
        setting['rpt_folder'] = ''  # 定義存放檢查報告的目錄
        setting['headers'] = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6,zh-CN;q=0.5,la;q=0.4',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Cookie': '_ga_SNR7NPLEYG=GS1.1.1651249603.1.0.1651250646.0; _ga_BGEHGPV3SB=GS1.1.1707539001.1.1.1707539323.0.0.0; _gid=GA1.3.1607128698.1707801742; _ga=GA1.1.381372473.1651249604; _ga_Q0EL30K2K5=GS1.1.1707801741.1.0.1707801770.0.0.0; _ga_54MVLT2EZN=GS1.1.1707843944.5.1.1707843969.0.0.0; __RequestVerificationToken_L05ldFNlcnZpY2Vz0=MrnqY4BqFXwyAR3uGWq5prQZPEwGWyzIJgIpuGFLyP8hqJ6eLKM9EWlC8NVA4MZqmyjtxmWT-9ZtzrO04NCTXcM_njimY7J0_WFWHlyWtzE1',
            'Sec-Ch-Ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        }
        setting['avoid_urls'] = [
            "https://accessibility.moda.gov.tw/Applications/DetectLog/157716",
            "https://accessibility.moda.gov.tw/Home/Info/",
            "https://accessibility.moda.gov.tw/Download/Detail/1375?Category=52",
            "https://www.ndc.gov.tw/cp.aspx?n=32A75A78342B669D]",
        ]
        setting['scan_urls'] = ["https://www.ncut.edu.tw/"]  # 定義想要檢查的網址
        yaml.dump(setting, f)
    return setting


def read_config(cfg_file) -> dict:
    # 讀取設定檔 config.yaml
    yaml = YAML()
    if os.path.exists(config_file):
        with open(cfg_file, 'r', encoding='UTF-8') as f:
            setting = yaml.load(f)
    else:
        threading.Thread(target=dl_resources).start()  # 下載相關元件
        setting = create_config(config_file)
        messagebox.showinfo("資訊", "首次執行，為您建立所需環境，請稍待...")
        while not os.path.exists('LocalVersion.yaml'):
            time.sleep(1)
    return setting


def make_cmd(file) -> None:
    '''產生更新程式 update.cmd'''
    with open(file, 'w') as f:
        f.write('@echo off\n')
        f.write('echo 進行新舊版執行檔替換作業...\n')
        f.write('if exist chklink.exe taskkill /f /im chklink.exe 2>nul\n')
        f.write('timeout 1\n')
        f.write('if exist chklink_upd.exe move /Y chklink.exe chklink.exe.old\n')
        f.write('if exist chklink_upd.exe move /Y chklink_upd.exe chklink.exe\n')
        f.write('if not exist chklink_upd.exe echo 更新成功！\n')
        f.write('if exist chklink_upd.exe echo 更新失敗！\n')
        f.write('start chklink.exe\n')
        f.write('timeout 6\n')


def update_setting(key, default_value):
    global updated
    if not setting.get(key):
        setting[key] = default_value
        updated = True


# 設定全域變數
HEADERS, AVOID_URLS, SCAN_URLS = dict(), list(), list()
stop_scan = False
upd_file = 'update.cmd'
if not os.path.exists(upd_file):
    make_cmd(upd_file)  # 建立更新程式的批次檔

config_file = 'config.yaml'
setting = read_config(config_file)  # 讀取設定檔 config.yaml，若設定檔存在則讀取，否則建立設定檔

# 檢查並補充缺少的設定
updated = False

update_setting('rpt_folder', os.path.join(os.environ['USERPROFILE'], 'Documents'))
if not os.path.exists(setting['rpt_folder']):
    setting['rpt_folder'] = os.path.join(os.environ['USERPROFILE'], 'Documents')
    updated = True
update_setting('check_http', 'yes')

# 如果有更新設定，則將更新後的設定存回設定檔
if updated:
    with open(config_file, 'w', encoding='UTF-8') as f:
        yaml = YAML()
        yaml.preserve_quotes = True
        yaml.indent(mapping=2, sequence=4, offset=2)
        yaml.dump(setting, f)


# 其它全域變數與設定
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # 關閉 SSL 不安全的警告訊息
logging.getLogger('requests').setLevel(logging.ERROR)  # 設定 requests 的 log 等級
logging.captureWarnings(True)  # 捕捉 py waring message
logger = None  # 定義 logger
visted_link_file = 'visited_link.yaml'  # 已檢查過的連結檔案
visited_link = load_visted_link(visted_link_file)  # 載入已檢查過的連結檔案，供儲存已檢查過的連結與回應狀態碼
browser = None  # 定義瀏覽器物件

# 建立主視窗
yaml = YAML()
vc = dict()
if os.path.exists('LocalVersion.yaml'):
    with open('LocalVersion.yaml', 'r', encoding='UTF-8') as f:
        vc = yaml.load(f)  # 讀取本地版本檔案

# 建立主視窗
form = ttk.Window(themename="superhero")
form.title(f"網頁失效連結掃描工具 Ver.{vc.get('version', "1.0")}")  # 設定視窗標題
form.geometry("1024x768")  # 設定視窗寬高
form.resizable(True, True)
# 指定行和列的權重
form.rowconfigure(0, weight=1)
form.columnconfigure(0, weight=1)

form.option_add("*Font", "新細明體 11")  # 設定所有元件的字型為新細明體，大小為 11

# default_font = font.nametofont("TkDefaultFont")  # 取得預設字型
# default_font.configure(family="Courier New", size=11)  # 設定預設字型為新細明體，大小為 11
# form.option_add("*Font", default_font)  # 設定所有元件的字型為預設字型

# 建立頁籤元件 Notebook
notebook = ttk.Notebook(form)
notebook.grid(row=0, column=0, padx=4, pady=4, sticky="nsew")

# 建立分頁
page1 = ttk.Frame(notebook)
page2 = ttk.Frame(notebook)
notebook.add(page1, text="網址掃描")
notebook.add(page2, text="系統設定")
page1.columnconfigure(0, weight=1)
page1.rowconfigure(2, weight=1)
page2.columnconfigure(0, weight=1)
page2.rowconfigure(3, weight=1)
page2.rowconfigure(4, weight=1)
page2.rowconfigure(5, weight=1)

# --- page1 ---
# 建立 frame1_1
frame1_1 = ttk.Labelframe(page1, text="URLs", style='success.TLabelframe')
frame1_1.grid(row=0, column=0, padx=5, pady=1, sticky="nsew")
frame1_1.columnconfigure(0, weight=1)

# 建立網址輸入框
start_url_txt = ttk.ScrolledText(frame1_1, wrap=tk.WORD, height=4, font=('Courier New', 10))
start_url_txt.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
start_url_txt.insert(1.0, ",".join(setting.get('scan_urls')) if setting.get('scan_urls') else "")
Hovertip(start_url_txt, '輸入想要掃描的網址，若有多個網址請以逗號隔開')

# 建立右鍵選單
start_url_context_menu = tk.Menu(start_url_txt, tearoff=0)
start_url_context_menu.add_command(label="剪下", command=lambda: start_url_txt.event_generate("<<Cut>>"))
start_url_context_menu.add_command(label="複製", command=lambda: start_url_txt.event_generate("<<Copy>>"))
start_url_context_menu.add_command(label="貼上", command=lambda: start_url_txt.event_generate("<<Paste>>"))
start_url_context_menu.add_separator()
start_url_context_menu.add_command(
    label="刪除",
    command=lambda: start_url_txt.delete(tk.SEL_FIRST, tk.SEL_LAST) if start_url_txt.get(1.0, tk.END) else None,
)
start_url_context_menu.add_command(label="清除", command=clear_start_url)
# 綁定右鍵事件
start_url_txt.bind("<Button-3>", lambda e: start_url_context_menu.post(e.x_root, e.y_root))
# bug fix: 修正 ttkbootstrap 無法在 ScrolledText 元件按下 Ctrl+A 選取所有文字的問題
# "d:\pyTest\chklink\.venv\Lib\site-packages\ttkbootstrap\window.py", line 104, in on_select_all
# if widget.__class__.__name__ in ("Text",'ScrolledText'):

# 建立掃描按鈕
analysis_btn = ttk.Button(
    frame1_1,
    text="掃描",
    command=lambda: analysis_func(start_url_txt.get(1.0, tk.END), int(layer_txt.get())),
    bootstyle="success",
    cursor='hand2',
)
Hovertip(analysis_btn, '按下按鈕開始掃描網址內的失效連結')
analysis_btn.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")

# 建立中斷按鈕
stop_btn = ttk.Button(
    frame1_1, text="中斷", command=stop_scanning, bootstyle="danger", cursor='X_cursor', state=tk.DISABLED
)
Hovertip(stop_btn, '按下按鈕中斷掃描')
stop_btn.grid(row=0, column=3, padx=5, pady=5, sticky="nsew")

# 建立 frame1_2
frame1_2 = ttk.Frame(page1)
frame1_2.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
check_var = tk.IntVar()
check_var.set(1)
check_button = ttk.Checkbutton(frame1_2, text="跳過已檢查過的網址", variable=check_var)
check_button.grid(row=0, column=0, padx=5, pady=5, sticky="e")
frame1_2.columnconfigure(0, weight=1)

# 建立 frame1_3
frame1_3 = ttk.Labelframe(page1, text="掃描歷史", style='success.TLabelframe')
frame1_3.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")
frame1_3.columnconfigure(0, weight=1)
frame1_3.rowconfigure(0, weight=1)

log_console = scrolledtext.ScrolledText(frame1_3, wrap=tk.WORD, font=('Courier New', 10))
log_console.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
log_console.tag_config('ERROR', foreground='#e64530')  # 設定 ERROR 的文字顏色
log_console.tag_config('INFO', foreground='#c5c6c4')  # 設定 INFO 的文字顏色
log_console.tag_config('WARNING', foreground='#ffcc00')  # 設定 WARNING 的文字顏色
log_console.tag_config('SUCCESS', foreground='#8ae234')  # 設定 SUCCESS 的文字顏色

# --- page2 ---
# 建立 frame2_1
frame2_1 = ttk.Frame(page2)
frame2_1.grid(row=0, column=0, padx=1, pady=1, sticky="nsew")
# 建立儲存設定按鈕
cfg_save_btn = ttk.Button(page2, text="儲存設定", command=save_config, bootstyle="info", cursor='hand2')
cfg_save_btn.grid(row=0, column=1, rowspan=2, padx=10, pady=10, sticky="nsew")
# 建立檢查更新按鈕
run_upd_btn = ttk.Button(page2, text="檢查更新", command=run_update, bootstyle="warning", cursor='hand2')
run_upd_btn.grid(row=0, column=2, rowspan=2, padx=10, pady=10, sticky="nsew")

# 建立檢查連結的層數Label
layer_txt_label = ttk.Label(frame2_1, text="檢查連結的層數")
layer_txt_label.grid(row=0, column=0, padx=5)

# 建立檢查連結的層數
layer_txt = ttk.Entry(frame2_1, width=3)
layer_txt.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
layer_txt.insert(0, setting.get('layer'))  # 設定預設連結的層數

# 建立圖片是否必須有 alt 屬性Label
alt_must_txt_label = ttk.Label(frame2_1, text="圖片是否必須有 alt 屬性")
alt_must_txt_label.grid(row=0, column=2, padx=5)

# 建立圖片是否必須有 alt 屬性
alt_must_txt = ttk.Combobox(frame2_1, values=["no", "yes"], width=6)
alt_must_txt.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
alt_must_txt.set(setting.get('alt_must'))

# 建立檢查不安全的 http 協定 Label
check_http_txt_label = ttk.Label(frame2_1, text="檢查不安全的 http 協定")
check_http_txt_label.grid(row=0, column=4, padx=5)

# 建立檢查不安全的 http 協定
check_http_txt = ttk.Combobox(frame2_1, values=["no", "yes"], width=6)
check_http_txt.grid(row=0, column=5, padx=5, pady=5, sticky="ew")
check_http_txt.set(setting.get('check_http'))

# 建立 frame2_2
frame2_2 = ttk.Frame(page2)
frame2_2.grid(row=1, column=0, padx=1, pady=1, sticky="nsew")
# frame2_2.columnconfigure(1, weight=1)
frame2_2.columnconfigure(4, weight=1)

# 建立連線逾時秒數Label
timeout_txt_label = ttk.Label(frame2_2, text="連線逾時秒數")
timeout_txt_label.grid(row=0, column=1, padx=5)

# 建立連線逾時秒數
timeout_txt = ttk.Entry(frame2_2, width=3)
timeout_txt.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
timeout_txt.insert(0, setting.get('timeout'))  # 設定預設連線逾時秒數

# 建立結果存放Label
report_dir_txt_label = ttk.Label(frame2_2, text="報告路徑")
report_dir_txt_label.grid(row=0, column=3, padx=5, pady=5)

# 建立結果存放輸入框
report_dir_txt = ttk.Entry(frame2_2, font=('Courier New', 10))
report_dir_txt.grid(row=0, column=4, padx=5, pady=5, sticky="ew")
report_dir_txt.insert(0, setting.get('rpt_folder'))  # 設定預設結果存放目錄

# 建立更改目錄按鈕
choose_dir_icon = tk.PhotoImage(file='icon\\folder.png')  # 設定更改目錄按鈕的圖示
choose_dir_button = ttk.Button(
    frame2_2,
    text="更改目錄",
    command=choose_dir,
    bootstyle="success-outline",
    cursor='hand2',
    image=choose_dir_icon,
    compound=tk.LEFT,  # 將圖示放在文字的左邊
)
choose_dir_button.grid(row=0, column=5, padx=10)

upd_progress_show = ttk.Progressbar(page2, orient='horizontal', mode='determinate')
upd_progress_show.grid(row=2, column=0, columnspan=3, padx=5, pady=2, sticky="nsew")

# 建立 frame2_3 用 LabelFrame
frame2_3 = ttk.Labelframe(page2, text="請求的標頭", style='info.TLabelframe')
frame2_3.grid(row=3, column=0, columnspan=3, padx=3, pady=5, sticky="nsew")
frame2_3.columnconfigure(0, weight=1)
frame2_3.rowconfigure(0, weight=1)

# 建立請求的標頭
headers_txt = scrolledtext.ScrolledText(frame2_3, wrap=tk.WORD, font=('Courier New', 10))
headers_txt.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
for header in setting.get('headers'):
    headers_txt.insert(tk.END, f"{header}: {setting.get('headers')[header]}\n")

# 建立 frame2_4 用 LabelFrame
frame2_4 = ttk.Labelframe(page2, text="想要避免檢查的網址清單", style='info.TLabelframe')
frame2_4.grid(row=4, column=0, columnspan=3, padx=3, pady=5, sticky="nsew")
frame2_4.columnconfigure(0, weight=1)
frame2_4.rowconfigure(0, weight=1)

# 建立想要避免檢查的網址清單
avoid_urls_txt = scrolledtext.ScrolledText(frame2_4, wrap=tk.WORD, font=('Courier New', 10))
avoid_urls_txt.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
for url in setting.get('avoid_urls'):
    avoid_urls_txt.insert(tk.END, f"{url}\n")

# 建立 frame2_5 用 LabelFrame
frame2_5 = ttk.Labelframe(page2, text="想要檢查的網址清單", style='info.TLabelframe')
frame2_5.grid(row=5, column=0, columnspan=3, padx=3, pady=5, sticky="nsew")
frame2_5.columnconfigure(0, weight=1)
frame2_5.rowconfigure(0, weight=1)

# 建立想要檢查的網址清單
scan_urls_txt = scrolledtext.ScrolledText(frame2_5, wrap=tk.WORD, font=('Courier New', 10))
scan_urls_txt.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
for url in setting.get('scan_urls'):
    scan_urls_txt.insert(tk.END, f"{url}\n")

# 開始主迴圈
form.mainloop()
